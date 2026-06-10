from datetime import date
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlmodel import Session, select

from app.database import engine, refresh_database
from app.main import app
from app.models import Doctor, DoctorTransaction, Treatment
from app.security import create_access_token


def auth_headers() -> dict[str, str]:
    return {"Authorization": f"Bearer {create_access_token('admin')}"}


def workbook_upload(sheet: str, headers: list[str], rows: list[list[object]]) -> BytesIO:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def seed_master() -> tuple[int, int]:
    refresh_database()
    with Session(engine) as session:
        doctor = Doctor(name="Drg. Dokter 1", normal_fee_rate=0.6)
        treatment = Treatment(name="Scaling A", bhp_cost=120_000, treatment_price=350_000)
        session.add(doctor)
        session.add(treatment)
        session.commit()
        session.refresh(doctor)
        session.refresh(treatment)
        return doctor.id, treatment.id


def test_transaction_create_edit_filter_and_delete():
    doctor_id, treatment_id = seed_master()
    client = TestClient(app)
    headers = auth_headers()

    response = client.post(
        "/doctor-transactions",
        headers=headers,
        json={
            "transaction_date": "2026-05-10",
            "doctor_id": doctor_id,
            "patient_name": "Pasien A",
            "treatment_id": treatment_id,
            "qty": 2,
            "discount_amount": 50_000,
            "needs_review": True,
            "review_note": "Cek ulang diskon",
        },
    )

    assert response.status_code == 200
    row = response.json()
    assert row["period"] == "2026-05"
    assert row["service_amount"] == 410_000
    assert row["doctor_fee_amount"] == 246_000
    assert row["total_bill_amount"] == 650_000
    assert row["needs_review"] is True
    assert row["review_note"] == "Cek ulang diskon"

    assert client.get("/doctor-transactions?period=2026-05", headers=headers).json()[0]["id"] == row["id"]
    assert client.get("/doctor-transactions?period=2026-06", headers=headers).json() == []

    update = client.patch(
        f"/doctor-transactions/{row['id']}",
        headers=headers,
        json={
            "period": "2026-05",
            "transaction_date": "2026-05-10",
            "doctor_id": doctor_id,
            "patient_name": "Pasien A",
            "treatment_id": treatment_id,
            "qty": 1,
            "discount_amount": 0,
            "needs_review": False,
            "review_note": "Sudah dicek",
        },
    )
    assert update.status_code == 200
    assert update.json()["doctor_fee_amount"] == 138_000
    assert update.json()["needs_review"] is False
    assert update.json()["review_note"] is None

    delete = client.delete(f"/doctor-transactions/{row['id']}", headers=headers)
    assert delete.status_code == 200
    assert client.get("/doctor-transactions?period=2026-05", headers=headers).json() == []


def test_transaction_import_preview_commit_marks_unknown_treatment_for_review():
    seed_master()
    client = TestClient(app)
    headers = auth_headers()
    upload = workbook_upload(
        "DoctorTransactions",
        [
            "period",
            "transaction_date",
            "doctor_name",
            "patient_name",
            "treatment_name",
            "qty",
            "discount_amount",
            "bhp_override",
            "price_override",
            "special_fee_amount",
            "fee_rate",
        ],
        [["2026-05", "2026-05-11", "Drg. Dokter 1", "Pasien B", "Treatment Baru", 1, 0, "", 100_000, 0, ""]],
    )

    preview_response = client.post(
        "/doctor-transactions/import/preview",
        headers=headers,
        files={"file": ("transactions.xlsx", upload)},
    )
    assert preview_response.status_code == 200
    preview = preview_response.json()
    assert preview["valid_rows"] == 1
    assert preview["summary"]["review"] == 1
    assert preview["rows"][0]["status"] == "review"

    commit = client.post(f"/doctor-transactions/import/{preview['import_id']}/commit", headers=headers)
    assert commit.status_code == 200
    assert commit.json()["created"] == 1

    with Session(engine) as session:
        trx = session.exec(select(DoctorTransaction)).one()
        assert trx.needs_review is True
        assert trx.period == "2026-05"


def test_doctor_transaction_template_includes_master_reference_notes():
    doctor_id, treatment_id = seed_master()
    client = TestClient(app)
    response = client.get("/reports/templates/doctor-transactions.xlsx")

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content), data_only=True)
    assert workbook.sheetnames == ["DoctorTransactions", "Notes"]

    notes = workbook["Notes"]
    values = [row for row in notes.iter_rows(values_only=True)]
    assert any(row[0] == "Referensi Dokter" for row in values)
    assert ["doctor_id", "doctor_name"] in [list(row[:2]) for row in values]
    assert [doctor_id, "Drg. Dokter 1"] in [list(row[:2]) for row in values]

    treatment_rows = [list(row[:11]) for row in values]
    assert [
        "treatment_id",
        "code",
        "treatment_name",
        "category",
        "doctor_cost",
        "specialist_cost",
        "bhp_cost",
        "service_fee",
        "treatment_price",
        "notes",
        "status",
    ] in treatment_rows
    assert any(row[0] == treatment_id and row[2] == "Scaling A" and row[6] == 120_000 and row[8] == 350_000 for row in treatment_rows)
