from datetime import date
from io import BytesIO
from zipfile import ZipFile

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlmodel import Session

from app.database import engine, refresh_database
from app.main import app
from app.models import Doctor, DoctorTransaction, Treatment
from app.security import create_access_token


def auth_headers() -> dict[str, str]:
    return {"Authorization": f"Bearer {create_access_token('admin')}"}


def seed_fee_rows(needs_review: bool = False) -> tuple[int, int]:
    refresh_database()
    with Session(engine) as session:
        doctor = Doctor(
            name="Drg. Leni Ruslaini",
            bank_name="MANDIRI",
            account_name="Drg. Leni Ruslaini",
            account_number="1234567890",
            nik="5201125809840001",
            normal_fee_rate=0.6,
            tax_rate=0.025,
        )
        treatment = Treatment(name="Scaling A", bhp_cost=120_000, treatment_price=350_000)
        session.add(doctor)
        session.add(treatment)
        session.commit()
        session.refresh(doctor)
        session.refresh(treatment)
        session.add(
            DoctorTransaction(
                period="2026-05",
                transaction_date=date(2026, 5, 10),
                doctor_id=doctor.id or 0,
                patient_name="Pasien A",
                treatment_id=None if needs_review else treatment.id,
                treatment_name_snapshot="Scaling A",
                qty=1,
                discount_amount=0,
                service_amount=230_000,
                doctor_fee_amount=138_000,
                total_bill_amount=350_000,
                needs_review=needs_review,
                review_note="Treatment belum ditemukan di master." if needs_review else None,
            )
        )
        session.commit()
        return doctor.id or 0, treatment.id or 0


def test_doctor_fee_overview_calculate_lock_and_transaction_guard():
    doctor_id, treatment_id = seed_fee_rows()
    client = TestClient(app)
    headers = auth_headers()

    overview = client.get("/doctor-periods/2026-05/overview", headers=headers)
    assert overview.status_code == 200
    assert overview.json()["status"] == "not_calculated"
    assert overview.json()["transaction_count"] == 1
    assert overview.json()["review_count"] == 0

    calculate = client.post("/doctor-periods/2026-05/calculate", headers=headers)
    assert calculate.status_code == 200

    overview = client.get("/doctor-periods/2026-05/overview", headers=headers).json()
    assert overview["status"] == "draft"
    assert overview["summaries"][0]["doctor_name"] == "Drg. Leni Ruslaini"
    assert overview["summaries"][0]["transaction_count"] == 1
    assert overview["summaries"][0]["account_number"] == "1234567890"
    assert overview["summaries"][0]["transfer_amount"] == 134_550

    lock = client.post("/doctor-periods/2026-05/lock", headers=headers)
    assert lock.status_code == 200
    assert lock.json()[0]["status"] == "locked"

    blocked_edit = client.post(
        "/doctor-transactions",
        headers=headers,
        json={
            "period": "2026-05",
            "transaction_date": "2026-05-12",
            "doctor_id": doctor_id,
            "patient_name": "Pasien B",
            "treatment_id": treatment_id,
        },
    )
    assert blocked_edit.status_code == 409
    assert client.post("/doctor-periods/2026-05/calculate", headers=headers).status_code == 409


def test_doctor_fee_lock_rejects_review_rows_and_export_has_detail_sheet():
    seed_fee_rows(needs_review=True)
    client = TestClient(app)
    headers = auth_headers()

    assert client.post("/doctor-periods/2026-05/calculate", headers=headers).status_code == 200
    lock = client.post("/doctor-periods/2026-05/lock", headers=headers)
    assert lock.status_code == 409
    assert "review" in lock.text

    response = client.get("/reports/doctor-fees?period=2026-05&format=xlsx", headers=headers)
    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content), data_only=True)
    assert workbook.sheetnames[0] == "Rekapan FEE DOKTER"
    assert "TS. Drg. Leni Ruslaini" in workbook.sheetnames
    assert workbook["Rekapan FEE DOKTER"]["N2"].value == "DRAFT"
    detail = workbook["TS. Drg. Leni Ruslaini"]
    assert detail["A1"].value == "DENTAL MANAGER"
    assert detail["B3"].value == "No. Surat: 001/SG-DOC/DENTAL MANAGER/V/2026"
    assert detail["A5"].value == "Tanggal"
    assert detail["C6"].value == "Scaling A"
    assert "A1:K1" in {str(item) for item in detail.merged_cells.ranges}

    pdf_response = client.get("/reports/doctor-fees?period=2026-05&format=pdf", headers=headers)
    assert pdf_response.status_code == 200
    assert pdf_response.headers["content-type"] == "application/pdf"
    assert pdf_response.content.startswith(b"%PDF")

    zip_response = client.get("/reports/doctor-fees?period=2026-05&format=zip", headers=headers)
    assert zip_response.status_code == 200
    assert zip_response.headers["content-type"] == "application/zip"
    archive = ZipFile(BytesIO(zip_response.content))
    names = archive.namelist()
    assert len(names) == 1
    assert names[0].endswith(".pdf")
    assert archive.read(names[0]).startswith(b"%PDF")

    archives = client.get("/reports/archive", headers=headers)
    assert archives.status_code == 200
    archive_rows = archives.json()
    filenames = {row["filename"] for row in archive_rows}
    assert "doctor-fees-2026-05.xlsx" in filenames
    assert "doctor-fees-2026-05.pdf" in filenames
    assert "doctor-fees-2026-05-per-dokter.zip" in filenames

    xlsx_archive = next(row for row in archive_rows if row["filename"] == "doctor-fees-2026-05.xlsx")
    archived_download = client.get(f"/reports/archive/{xlsx_archive['id']}/download", headers=headers)
    assert archived_download.status_code == 200
    assert archived_download.content.startswith(b"PK")

    deleted = client.delete(f"/reports/archive/{xlsx_archive['id']}", headers=headers)
    assert deleted.status_code == 200
    assert client.get(f"/reports/archive/{xlsx_archive['id']}/download", headers=headers).status_code == 404
