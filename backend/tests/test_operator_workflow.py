from datetime import date, datetime, timedelta, time
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlmodel import Session, select

from app.database import engine, prune_old_audit_logs, refresh_database
from app.main import app
from app.models import AuditLog, AttendanceRecord, Employee, User, UserRole
from app.security import create_access_token, hash_password


def admin_headers() -> dict[str, str]:
    return {"Authorization": f"Bearer {create_access_token('admin')}"}


def operator_headers() -> dict[str, str]:
    return {"Authorization": f"Bearer {create_access_token('operator')}"}


def seed_operator_data() -> tuple[int, int, int, int]:
    refresh_database()
    with Session(engine) as session:
        mine = Employee(
            name="Operator Satu",
            attendance_id="OP-1",
            position="Front Office",
            base_salary=2_500_000,
            working_days=25,
            bank_name="BCA",
            account_name="Operator Satu",
            account_number="111",
        )
        other = Employee(
            name="Operator Dua",
            attendance_id="OP-2",
            position="Perawat",
            base_salary=2_800_000,
            working_days=25,
            bank_name="BNI",
            account_name="Operator Dua",
            account_number="222",
        )
        session.add(mine)
        session.add(other)
        session.flush()
        mine_attendance = AttendanceRecord(
            period="2026-05",
            employee_id=mine.id,
            attendance_id_snapshot="OP-1",
            employee_name_snapshot=mine.name,
            work_date=date(2026, 5, 6),
            timezone1_in=time(8, 10),
            timezone1_out=time(17, 0),
            overtime_minutes=60,
        )
        other_attendance = AttendanceRecord(
            period="2026-05",
            employee_id=other.id,
            attendance_id_snapshot="OP-2",
            employee_name_snapshot=other.name,
            work_date=date(2026, 5, 6),
            timezone1_in=time(8, 0),
            timezone1_out=time(16, 0),
        )
        session.add(mine_attendance)
        session.add(other_attendance)
        session.add(
            User(
                username="operator",
                full_name="Operator Satu",
                role=UserRole.OPERATOR,
                employee_id=mine.id,
                hashed_password=hash_password("secret"),
                is_active=True,
            )
        )
        session.commit()
        return mine.id or 0, other.id or 0, mine_attendance.id or 0, other_attendance.id or 0


def test_operator_attendance_is_scoped_and_protest_flags_review():
    mine_id, _, mine_attendance_id, other_attendance_id = seed_operator_data()
    client = TestClient(app)

    rows = client.get("/attendance-records?period=2026-05", headers=operator_headers())
    assert rows.status_code == 200
    body = rows.json()
    assert len(body) == 1
    assert body[0]["employee_id"] == mine_id

    employees = client.get("/employees", headers=operator_headers())
    assert employees.status_code == 200
    assert [employee["id"] for employee in employees.json()] == [mine_id]

    blocked = client.post(
        f"/attendance-records/{other_attendance_id}/protest",
        headers=operator_headers(),
        json={"reason": "Jam masuk milik karyawan lain."},
    )
    assert blocked.status_code == 403

    protest = client.post(
        f"/attendance-records/{mine_attendance_id}/protest",
        headers=operator_headers(),
        json={"reason": "Fingerprint masuk tidak terbaca dengan tepat."},
    )
    assert protest.status_code == 200
    assert protest.json()["needs_review"] is True
    assert protest.json()["protest_note"] == "Fingerprint masuk tidak terbaca dengan tepat."

    admin_rows = client.get("/attendance-records?period=2026-05&review=true", headers=admin_headers())
    assert admin_rows.status_code == 200
    assert admin_rows.json()[0]["protest_by_name"] == "Operator Satu"


def test_operator_payroll_dashboard_and_exports_are_self_scoped():
    mine_id, _, _, _ = seed_operator_data()
    client = TestClient(app)
    assert client.post("/payroll-periods/2026-05/calculate", headers=admin_headers()).status_code == 200

    dashboard = client.get("/me/dashboard?period=2026-05", headers=operator_headers())
    assert dashboard.status_code == 200
    assert dashboard.json()["employee"]["id"] == mine_id
    assert dashboard.json()["attendance_count"] == 1

    payroll = client.get("/me/payroll/2026-05", headers=operator_headers())
    assert payroll.status_code == 200
    assert payroll.json()["payroll"]["employee_id"] == mine_id
    assert payroll.json()["payroll"]["net_salary"] > 0

    xlsx = client.get("/me/payroll/2026-05/export?format=xlsx", headers=operator_headers())
    assert xlsx.status_code == 200
    workbook = load_workbook(BytesIO(xlsx.content))
    names = [row[1].value for row in workbook["Form Gaji Karyawan"].iter_rows(min_row=4, max_col=2)]
    assert "Operator Satu" in names
    assert "Operator Dua" not in names

    pdf = client.get("/me/payroll/2026-05/export?format=pdf", headers=operator_headers())
    assert pdf.status_code == 200
    assert pdf.content.startswith(b"%PDF")


def test_login_logout_are_written_to_audit_log():
    seed_operator_data()
    client = TestClient(app)

    login = client.post("/auth/login", data={"username": "operator", "password": "secret"})
    assert login.status_code == 200
    token = login.json()["access_token"]
    logout = client.post("/auth/logout", headers={"Authorization": f"Bearer {token}"})
    assert logout.status_code == 200

    with Session(engine) as session:
        actions = session.exec(select(AuditLog).where(AuditLog.actor_username == "operator")).all()
    assert [row.action for row in actions] == ["login", "logout"]


def test_operator_audit_logs_are_self_scoped():
    seed_operator_data()
    client = TestClient(app)

    with Session(engine) as session:
        operator = session.exec(select(User).where(User.username == "operator")).one()
        session.add(
            AuditLog(
                actor_id=operator.id,
                actor_username=operator.username,
                actor_name=operator.full_name,
                action="login",
                entity_type="auth",
                description="Login operator.",
            )
        )
        session.add(
            AuditLog(
                actor_username="admin",
                actor_name="Administrator",
                action="delete",
                entity_type="doctor_transaction",
                description="Menghapus transaksi admin.",
            )
        )
        session.commit()

    response = client.get("/audit-logs/me?limit=50", headers=operator_headers())

    assert response.status_code == 200
    body = response.json()
    assert [row["actor_username"] for row in body] == ["operator"]
    assert body[0]["description"] == "Login operator."
    assert body[0]["created_at"].endswith("Z")


def test_old_audit_logs_are_pruned_automatically():
    refresh_database()
    with Session(engine) as session:
        session.add(
            AuditLog(
                actor_username="operator",
                actor_name="Operator Satu",
                action="login",
                entity_type="auth",
                description="Log lama.",
                created_at=datetime.utcnow() - timedelta(days=366),
            )
        )
        session.add(
            AuditLog(
                actor_username="operator",
                actor_name="Operator Satu",
                action="logout",
                entity_type="auth",
                description="Log baru.",
                created_at=datetime.utcnow() - timedelta(days=30),
            )
        )
        session.commit()

    assert prune_old_audit_logs() == 1

    with Session(engine) as session:
        rows = session.exec(select(AuditLog).order_by(AuditLog.created_at)).all()

    assert [row.description for row in rows] == ["Log baru."]
