from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlmodel import Session, select

from app.database import engine, refresh_database
from app.main import app
from app.models import Employee, PayrollRecord, PeriodStatus
from app.reports import template_xlsx
from app.security import create_access_token


def auth_headers() -> dict[str, str]:
    return {"Authorization": f"Bearer {create_access_token('admin')}"}


def attendance_workbook() -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    ws["A1"] = "Exc. Lap Statistik"
    ws["A3"] = "ID"
    ws["B3"] = "Nama"
    ws["C3"] = "Tgl"
    ws["D3"] = "Timezone I"
    ws["E3"] = "Timezone I"
    ws["F3"] = "Timezone II"
    ws["G3"] = "Timezone II"
    ws["H3"] = "Terlambat (Min)"
    ws["I3"] = "Pulang Awal (Min)"
    ws["J3"] = "Absen (Min)"
    ws["K3"] = "Total (Min)"
    ws["L3"] = "Libur"
    ws["M3"] = "Catatan"
    ws["D4"] = "Masuk"
    ws["E4"] = "Keluar"
    ws["F4"] = "Masuk"
    ws["G4"] = "Keluar"
    ws.append([1, "RI", "2026-05-02", "08:10", "16:20", None, None, None, None, None, None, "", ""])
    ws.append([99, "TIDAK ADA", "2026-05-03", "08:00", "16:00", None, None, None, None, None, None, "", "cek"])
    ws.append([1, "RI", "2026-05-04", None, None, None, None, None, None, None, None, "", "tidak hadir"])
    ws.append([1, "RI", "2026-05-03", None, None, None, None, None, None, None, None, "", "minggu kosong"])
    ws.append([1, "RI", "2026-05-10", "09:00", "12:30", None, None, None, None, None, None, "", "minggu masuk"])
    ws.append([1, "RI", "2026-05-06", "10:00", "13:00", None, None, None, None, None, None, "ya", "tanggal merah"])
    ws.append([1, "RI", "2026-05-17", "08:30", "15:30", None, None, None, None, None, None, "masuk", "minggu jadwal masuk"])
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def seed_employee() -> None:
    refresh_database()
    with Session(engine) as session:
        session.add(Employee(name="RIKA", attendance_id="1", base_salary=2_500_000, working_days=25))
        session.commit()


def test_attendance_preview_commit_upsert_and_lock_guard():
    seed_employee()
    client = TestClient(app)
    headers = auth_headers()

    workbook = attendance_workbook()
    preview = client.post(
        "/attendance/import-preview",
        headers=headers,
        files={"file": ("attendance.xlsx", workbook.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert preview.status_code == 200
    body = preview.json()
    assert body["valid_rows"] == 7
    assert body["summary"]["review"] == 1
    assert body["rows"][0]["status"] == "new"
    assert body["rows"][0]["late_minutes"] == 10
    assert body["rows"][0]["early_leave_minutes"] == 0
    assert body["rows"][0]["is_absent"] is False
    assert body["rows"][0]["total_minutes"] == 10
    assert body["rows"][0]["overtime_minutes"] == 20
    assert body["rows"][1]["status"] == "review"
    assert body["rows"][2]["is_absent"] is True
    assert body["rows"][2]["absent_minutes"] == 480
    assert body["rows"][2]["total_minutes"] == 0
    assert body["rows"][3]["is_absent"] is False
    assert body["rows"][3]["is_sunday"] is True
    assert body["rows"][3]["total_minutes"] == 0
    assert body["rows"][4]["is_absent"] is False
    assert body["rows"][4]["is_sunday"] is True
    assert body["rows"][4]["overtime_minutes"] == 210
    assert body["rows"][5]["is_holiday"] is True
    assert body["rows"][5]["overtime_minutes"] == 180
    assert body["rows"][6]["is_holiday"] is False
    assert body["rows"][6]["late_minutes"] == 30
    assert body["rows"][6]["early_leave_minutes"] == 30
    assert body["rows"][6]["overtime_minutes"] == 0

    commit = client.post(
        "/attendance/import",
        headers=headers,
        files={"file": ("attendance.xlsx", workbook.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert commit.status_code == 200
    assert commit.json()["created"] == 7

    second_commit = client.post(
        "/attendance/import",
        headers=headers,
        files={"file": ("attendance.xlsx", workbook.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert second_commit.status_code == 200
    assert second_commit.json()["updated"] == 7

    rows = client.get("/attendance-records?period=2026-05", headers=headers)
    assert rows.status_code == 200
    assert len(rows.json()) == 7
    assert rows.json()[0]["attendance_id_snapshot"] == "1"
    assert rows.json()[0]["employee_name_snapshot"] == "RI"
    assert rows.json()[0]["is_absent"] is False
    assert rows.json()[0]["overtime_minutes"] == 20
    absent_row = next(row for row in rows.json() if row["work_date"] == "2026-05-04")
    assert absent_row["is_absent"] is True
    assert absent_row["total_minutes"] == 0
    sunday_absent = next(row for row in rows.json() if row["work_date"] == "2026-05-03" and row["attendance_id_snapshot"] == "1")
    assert sunday_absent["is_absent"] is False
    assert sunday_absent["is_sunday"] is True
    assert sunday_absent["overtime_minutes"] == 0
    sunday_overtime = next(row for row in rows.json() if row["work_date"] == "2026-05-10")
    assert sunday_overtime["is_sunday"] is True
    assert sunday_overtime["overtime_minutes"] == 210
    holiday_overtime = next(row for row in rows.json() if row["work_date"] == "2026-05-06")
    assert holiday_overtime["is_holiday"] is True
    assert holiday_overtime["overtime_minutes"] == 180
    sunday_regular = next(row for row in rows.json() if row["work_date"] == "2026-05-17")
    assert sunday_regular["is_sunday"] is True
    assert sunday_regular["is_holiday"] is False
    assert sunday_regular["total_minutes"] == 60

    with Session(engine) as session:
        employee = session.exec(select(Employee).where(Employee.attendance_id == "1")).first()
        session.add(PayrollRecord(period="2026-05", employee_id=employee.id, status=PeriodStatus.LOCKED))
        session.commit()

    blocked = client.post(
        "/attendance/import",
        headers=headers,
        files={"file": ("attendance.xlsx", workbook.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert blocked.status_code == 409


def test_attendance_template_notes_include_employee_attendance_ids():
    seed_employee()
    with Session(engine) as session:
        session.add(Employee(name="TANPA ID", base_salary=1_000_000))
        session.commit()
        fallback_employee = session.exec(select(Employee).where(Employee.name == "TANPA ID")).first()
        workbook = template_xlsx("attendance", session)

    from openpyxl import load_workbook

    wb = load_workbook(workbook)
    notes = wb["Notes"]
    values = list(notes.values)
    assert ("attendance_id", "name", "status") in values
    assert ("1", "RIKA", "aktif") in values
    assert (str(fallback_employee.id), "TANPA ID", "aktif") in values
