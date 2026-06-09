from datetime import date, time
from io import BytesIO
from zipfile import ZipFile

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlmodel import Session
from sqlmodel import select

from app.database import engine, refresh_database
from app.main import app
from app.models import AttendanceRecord, Employee, PayrollRecord, PayrollRule
from app.security import create_access_token


def auth_headers() -> dict[str, str]:
    return {"Authorization": f"Bearer {create_access_token('admin')}"}


def seed_payroll_data(needs_review: bool = False) -> None:
    refresh_database()
    with Session(engine) as session:
        employee = Employee(
            name="Rika",
            attendance_id="1",
            position="Supervisor",
            base_salary=2_712_250,
            working_days=25,
            bank_name="BSI",
            account_name="Rika",
            account_number="123",
        )
        session.add(employee)
        session.flush()
        session.add(
            AttendanceRecord(
                period="2026-05",
                employee_id=employee.id,
                attendance_id_snapshot="1",
                employee_name_snapshot="RIKA",
                work_date=date(2026, 5, 5),
                timezone1_in=time(7, 50),
                timezone1_out=time(17, 44),
                overtime_minutes=104,
                status_note="lembur",
                needs_review=needs_review,
            )
        )
        session.commit()


def test_payroll_overview_adjustment_export_and_lock_policy():
    seed_payroll_data()
    client = TestClient(app)
    headers = auth_headers()

    initial_overview = client.get("/payroll-periods/2026-05/overview", headers=headers)
    assert initial_overview.status_code == 200
    initial_body = initial_overview.json()
    assert initial_body["status"] == "not_calculated"
    assert initial_body["employee_count"] == 1
    assert initial_body["summaries"][0]["id"] is None
    assert initial_body["summaries"][0]["employee_name"] == "Rika"

    calculate = client.post("/payroll-periods/2026-05/calculate", headers=headers)
    assert calculate.status_code == 200

    overview = client.get("/payroll-periods/2026-05/overview", headers=headers)
    assert overview.status_code == 200
    body = overview.json()
    assert body["status"] == "draft"
    assert body["employee_count"] == 1
    assert body["total_overtime_minutes"] == 104
    payroll_id = body["summaries"][0]["id"]

    patched = client.patch(
        f"/payroll-records/{payroll_id}",
        headers=headers,
        json={
            "bonus": 100000,
            "position_allowance": 50000,
            "other_deduction": 25000,
            "izin_count": 1,
            "sakit_count": 0,
            "cuti_count": 0,
            "alpha_count": 0,
            "payment_method": "Transfer",
            "bank_name": "BSI",
            "account_name": "Rika",
            "account_number": "999",
            "needs_review": False,
        },
    )
    assert patched.status_code == 200
    assert patched.json()["bonus"] == 100000
    assert patched.json()["account_number"] == "999"

    overtime = client.get("/payroll-periods/2026-05/overtime?employee_id=1", headers=headers)
    assert overtime.status_code == 200
    assert overtime.json()[0]["overtime_minutes"] == 104

    xlsx = client.get("/reports/payroll?period=2026-05&format=xlsx", headers=headers)
    assert xlsx.status_code == 200
    wb = load_workbook(BytesIO(xlsx.content))
    assert {"Form Gaji Karyawan", "REKAPAN LEMBUR", "SLIP GAJI"}.issubset(set(wb.sheetnames))

    pdf = client.get("/reports/payroll?period=2026-05&format=pdf", headers=headers)
    assert pdf.status_code == 200
    assert pdf.content.startswith(b"%PDF")

    zip_response = client.get("/reports/payroll?period=2026-05&format=zip", headers=headers)
    assert zip_response.status_code == 200
    with ZipFile(BytesIO(zip_response.content)) as archive:
        assert archive.namelist()
        assert archive.read(archive.namelist()[0]).startswith(b"%PDF")

    lock = client.post("/payroll-periods/2026-05/lock", headers=headers)
    assert lock.status_code == 200


def test_payroll_lock_fails_when_attendance_needs_review():
    seed_payroll_data(needs_review=True)
    client = TestClient(app)
    headers = auth_headers()

    assert client.post("/payroll-periods/2026-05/calculate", headers=headers).status_code == 200
    blocked = client.post("/payroll-periods/2026-05/lock", headers=headers)
    assert blocked.status_code == 409


def test_payroll_recalculate_refreshes_overtime_from_attendance():
    seed_payroll_data()
    with Session(engine) as session:
        employee = session.exec(select(Employee).where(Employee.attendance_id == "1")).first()
        session.add(
            PayrollRecord(
                period="2026-05",
                employee_id=employee.id,
                base_salary=employee.base_salary,
                working_days=employee.working_days,
                overtime_minutes=0,
                overtime_total=0,
            )
        )
        session.commit()

    client = TestClient(app)
    headers = auth_headers()
    calculate = client.post("/payroll-periods/2026-05/calculate", headers=headers)
    assert calculate.status_code == 200

    overview = client.get("/payroll-periods/2026-05/overview", headers=headers)
    assert overview.status_code == 200
    summary = overview.json()["summaries"][0]
    assert summary["overtime_minutes"] == 104
    assert summary["overtime_total"] == 104 * summary["overtime_rate_per_minute"]


def test_payroll_uses_default_base_salary_and_training_reduction():
    refresh_database()
    with Session(engine) as session:
        rule = session.exec(select(PayrollRule).where(PayrollRule.is_default == True)).first()  # noqa: E712
        rule.default_base_salary = 3_000_000
        session.add(rule)
        session.add(
            Employee(
                name="Training",
                attendance_id="10",
                position="Perawat",
                base_salary=0,
                working_days=25,
                is_training=True,
            )
        )
        session.commit()

    client = TestClient(app)
    headers = auth_headers()
    calculate = client.post("/payroll-periods/2026-05/calculate", headers=headers)
    assert calculate.status_code == 200

    overview = client.get("/payroll-periods/2026-05/overview", headers=headers)
    assert overview.status_code == 200
    summary = overview.json()["summaries"][0]
    assert summary["base_salary"] == 2_400_000
    assert summary["is_training"] is True
