from io import BytesIO
import re
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.platypus import Paragraph, Table, TableStyle
from sqlmodel import Session, select

from app.config import get_settings
from app.models import AttendanceRecord, Doctor, DoctorFeeRule, DoctorPeriodSummary, DoctorTransaction, Employee, PayrollRecord, PeriodStatus, Treatment


def _money(value: float) -> str:
    return f"Rp {value:,.0f}".replace(",", ".")


def _number(value: float | int | None) -> str:
    if value in (None, ""):
        return "-"
    return f"{float(value):,.0f}".replace(",", ".")


def _rate(value: float | None) -> str:
    return f"{(value or 0) * 100:.0f}%"


def _clinic_name() -> str:
    return get_settings().app_name.upper()


def _roman_month(period: str) -> str:
    month = int(period.split("-")[1])
    return ["I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X", "XI", "XII"][month - 1]


def _period_year(period: str) -> str:
    return period.split("-")[0]


def _period_label(period: str) -> str:
    month = int(period.split("-")[1])
    names = ["Januari", "Februari", "Maret", "April", "Mei", "Juni", "Juli", "Agustus", "September", "Oktober", "November", "Desember"]
    return f"{names[month - 1]} {_period_year(period)}"


def _letter_number(index: int, period: str) -> str:
    return f"{index:03d}/SG-DOC/{_clinic_name()}/{_roman_month(period)}/{_period_year(period)}"


def _safe_sheet_title(title: str, used: set[str]) -> str:
    base = re.sub(r"[:\\/?*\[\]]", " ", title).strip() or "Sheet"
    base = base[:31]
    candidate = base
    index = 2
    while candidate in used:
        suffix = f" {index}"
        candidate = f"{base[:31 - len(suffix)]}{suffix}"
        index += 1
    used.add(candidate)
    return candidate


def _safe_filename(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9._ -]+", " ", value).strip().replace("  ", " ") or "file"


thin_black = Side(style="thin", color="000000")
thin_gray = Side(style="thin", color="D7DEE8")
table_border = Border(left=thin_black, right=thin_black, top=thin_black, bottom=thin_black)
soft_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)


def _autosize(ws, minimum: int = 10, maximum: int = 34) -> None:
    for column in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        ws.column_dimensions[get_column_letter(column[0].column)].width = min(max(max_length + 2, minimum), maximum)


def _style_summary_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    total_fill = PatternFill("solid", fgColor="D9EAF7")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center")
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'
            cell.border = soft_border
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    if ws.max_row > 1:
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
            cell.fill = total_fill
    ws.freeze_panes = "A2"
    _autosize(ws, 12, 32)


def _style_detail_sheet(ws, table_start: int, table_end: int, total_row: int) -> None:
    title_fill = PatternFill("solid", fgColor="4F81BD")
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    total_fill = PatternFill("solid", fgColor="F4B183")
    for row in range(1, ws.max_row + 1):
        for col in range(1, 12):
            cell = ws.cell(row, col)
            cell.alignment = Alignment(vertical="center")
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'

    for row in (1, 2):
        for col in range(1, 12):
            cell = ws.cell(row, col)
            cell.fill = title_fill
            cell.font = Font(bold=row == 1, color="000000", size=12 if row == 1 else 11)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[table_start]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.border = table_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in range(table_start + 1, table_end + 1):
        for col in range(1, 12):
            ws.cell(row, col).border = table_border
        for col in (4, 5, 7, 8, 9, 10, 11):
            ws.cell(row, col).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row, 6).alignment = Alignment(horizontal="right", vertical="center")

    for col in range(4, 12):
        cell = ws.cell(total_row, col)
        cell.border = table_border
        cell.fill = total_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(total_row, 4).alignment = Alignment(horizontal="center", vertical="center")

    widths = [12, 22, 40, 14, 16, 8, 14, 16, 16, 18, 22]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = f"A{table_start + 1}"


def _draft_summary(session: Session, period: str) -> list[DoctorPeriodSummary]:
    default_rule = session.exec(select(DoctorFeeRule).where(DoctorFeeRule.is_default == True)).first()  # noqa: E712
    default_rule = default_rule or DoctorFeeRule(name="Default", is_default=True)
    transactions = session.exec(select(DoctorTransaction).where(DoctorTransaction.period == period)).all()
    by_doctor: dict[int, list[DoctorTransaction]] = {}
    for transaction in transactions:
        by_doctor.setdefault(transaction.doctor_id, []).append(transaction)

    summaries: list[DoctorPeriodSummary] = []
    for doctor_id, rows in by_doctor.items():
        doctor = session.get(Doctor, doctor_id)
        tax_rate = doctor.tax_rate if doctor else default_rule.tax_rate
        treatment_fee = sum(row.doctor_fee_amount for row in rows if not row.special_fee_amount)
        ortho_fee = sum(row.special_fee_amount for row in rows)
        total_fee = treatment_fee + ortho_fee
        total_bill = sum(row.total_bill_amount for row in rows)
        tax = total_fee * tax_rate
        summaries.append(
            DoctorPeriodSummary(
                period=period,
                doctor_id=doctor_id,
                status=PeriodStatus.DRAFT,
                treatment_fee_total=treatment_fee,
                ortho_fee_total=ortho_fee,
                total_fee=total_fee,
                total_bill=total_bill,
                deduction=default_rule.default_deduction,
                tax=tax,
                transfer_amount=total_fee - default_rule.default_deduction - tax,
            )
        )
    return summaries


def _transaction_amounts(session: Session, transaction: DoctorTransaction) -> tuple[float, float]:
    treatment = session.get(Treatment, transaction.treatment_id) if transaction.treatment_id else None
    bhp = transaction.bhp_override if transaction.bhp_override is not None else (treatment.bhp_cost if treatment else 0)
    price = transaction.price_override if transaction.price_override is not None else (treatment.treatment_price if treatment else 0)
    return bhp, price


def doctor_fee_xlsx(session: Session, period: str) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Rekapan FEE DOKTER"
    summaries, export_status = _doctor_fee_summaries(session, period)

    ws.append(["NO", "NIP", "NAMA", "FEE DOKTER", "FEE ORTHODONTI", "TOTAL FEE DOKTER", "TOTAL BILL PASIEN", "Potongan dokter", "Pajak", "NOMINAL TRANSFER", "BANK TRANSFER", "NO REKENING", "PERIODE", "STATUS"])
    for index, summary in enumerate(summaries, start=1):
        doctor = session.get(Doctor, summary.doctor_id)
        ws.append(
            [
                index,
                doctor.nik if doctor else None,
                doctor.name if doctor else summary.doctor_id,
                summary.treatment_fee_total,
                summary.ortho_fee_total,
                summary.total_fee,
                summary.total_bill,
                summary.deduction,
                summary.tax,
                summary.transfer_amount,
                doctor.bank_name if doctor else None,
                doctor.account_number if doctor else None,
                _period_label(period),
                export_status,
            ]
        )
    if summaries:
        ws.append(
            [
                None,
                None,
                "TOTAL",
                sum(row.treatment_fee_total for row in summaries),
                sum(row.ortho_fee_total for row in summaries),
                sum(row.total_fee for row in summaries),
                sum(row.total_bill for row in summaries),
                sum(row.deduction for row in summaries),
                sum(row.tax for row in summaries),
                sum(row.transfer_amount for row in summaries),
                None,
                None,
                _period_label(period),
                export_status,
            ]
        )
    _style_summary_sheet(ws)

    used_titles = {ws.title}
    for index, summary in enumerate(summaries, start=1):
        doctor = session.get(Doctor, summary.doctor_id)
        doctor_name = doctor.name if doctor else f"Dokter {summary.doctor_id}"
        detail = wb.create_sheet(_safe_sheet_title(f"TS. {doctor_name}", used_titles))
        detail.merge_cells("A1:K1")
        detail.merge_cells("A2:K2")
        detail["A1"] = _clinic_name()
        detail["A2"] = "SLIP PENDAPATAN DOKTER"
        detail["B3"] = f"No. Surat: {_letter_number(index, period)}"
        detail["J3"] = export_status
        table_start = 5
        detail.append([])
        detail.append(["Tanggal", "Nama Pasien", "Perawatan", "BHP", "Biaya Perawatan", "QTY", "Diskon", "BIAYA JASA", "FEE DOKTER", "FEE KHUSUS BEHEL", "TOTAL BIAYA PERAWATAN"])
        rows = session.exec(
            select(DoctorTransaction).where(DoctorTransaction.period == period, DoctorTransaction.doctor_id == summary.doctor_id)
        ).all()
        rows = sorted(rows, key=lambda item: (item.transaction_date, item.patient_name.casefold(), item.id or 0))
        first_row_by_date: dict[object, int] = {}
        last_row_by_date: dict[object, int] = {}
        first_row_by_patient: dict[tuple[object, str], int] = {}
        last_row_by_patient: dict[tuple[object, str], int] = {}
        for transaction in rows:
            bhp, price = _transaction_amounts(session, transaction)
            current_row = detail.max_row + 1
            first_row_by_date.setdefault(transaction.transaction_date, current_row)
            last_row_by_date[transaction.transaction_date] = current_row
            patient_key = (transaction.transaction_date, transaction.patient_name.casefold())
            first_row_by_patient.setdefault(patient_key, current_row)
            last_row_by_patient[patient_key] = current_row
            detail.append(
                [
                    transaction.transaction_date,
                    transaction.patient_name,
                    transaction.treatment_name_snapshot,
                    bhp,
                    price,
                    transaction.qty,
                    transaction.discount_amount,
                    transaction.service_amount,
                    transaction.doctor_fee_amount,
                    transaction.special_fee_amount,
                    transaction.total_bill_amount,
                ]
            )
        for row_date, start_row in first_row_by_date.items():
            end_row = last_row_by_date[row_date]
            if end_row > start_row:
                detail.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
        for patient_key, start_row in first_row_by_patient.items():
            end_row = last_row_by_patient[patient_key]
            if end_row > start_row:
                detail.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
        detail.append([])
        total_row = detail.max_row + 1
        detail.merge_cells(start_row=total_row, start_column=4, end_row=total_row, end_column=7)
        detail.cell(total_row, 4).value = f"TOTAL FEE {doctor_name}"
        detail.cell(total_row, 8).value = f"FEE PERAWATAN {_rate(doctor.normal_fee_rate if doctor else None)}"
        detail.cell(total_row, 9).value = summary.treatment_fee_total
        detail.cell(total_row, 10).value = f"FEE ORTHO {_rate(doctor.ortho_fee_rate if doctor else None)}"
        detail.cell(total_row, 11).value = summary.ortho_fee_total
        detail.append([None, None, None, None, None, None, None, "Potongan", summary.deduction, "Pajak", summary.tax])
        detail.append([None, None, None, None, None, None, None, "Nominal Transfer", None, None, summary.transfer_amount])
        _style_detail_sheet(detail, table_start, table_start + len(rows), total_row)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def _doctor_fee_summaries(session: Session, period: str) -> tuple[list[DoctorPeriodSummary], str]:
    summaries = session.exec(select(DoctorPeriodSummary).where(DoctorPeriodSummary.period == period)).all()
    if not summaries:
        summaries = _draft_summary(session, period)
    export_status = "FINAL" if any(summary.status == PeriodStatus.LOCKED for summary in summaries) else "DRAFT"
    return sorted(summaries, key=lambda item: item.doctor_id), export_status


def _draw_pdf_header(pdf: canvas.Canvas, title: str, period: str, status: str) -> float:
    width, height = landscape(A4)
    pdf.setFont("Helvetica-Bold", 15)
    pdf.drawString(16 * mm, height - 18 * mm, title)
    pdf.setFont("Helvetica", 9)
    status_text = f" | {status}" if status == "DRAFT" else ""
    pdf.drawRightString(width - 16 * mm, height - 18 * mm, f"Periode {period}{status_text}")
    pdf.line(16 * mm, height - 23 * mm, width - 16 * mm, height - 23 * mm)
    return height - 32 * mm


def _draw_key_values(pdf: canvas.Canvas, rows: list[tuple[str, str]], x: float, y: float) -> float:
    pdf.setFont("Helvetica", 9)
    for label, value in rows:
        pdf.drawString(x, y, label)
        pdf.drawRightString(x + 62 * mm, y, value)
        y -= 6 * mm
    return y


def _short_text(value: str, limit: int) -> str:
    return value if len(value) <= limit else f"{value[: max(0, limit - 1)]}..."


def doctor_fee_pdf(session: Session, period: str, doctor_id: int | None = None, include_summary: bool = True) -> BytesIO:
    all_summaries, export_status = _doctor_fee_summaries(session, period)
    summaries = [summary for summary in all_summaries if doctor_id is None or summary.doctor_id == doctor_id]
    stream = BytesIO()
    pdf = canvas.Canvas(stream, pagesize=landscape(A4))
    width, height = landscape(A4)
    left = 10 * mm
    right = width - 10 * mm

    if include_summary:
        y = _draw_pdf_header(pdf, "REKAPAN FEE DOKTER", period, export_status)

        summary_data = [["NO", "NAMA", "FEE DOKTER", "FEE ORTHO", "TOTAL BILL", "PAJAK", "TRANSFER", "BANK", "NO REKENING"]]
        for index, summary in enumerate(all_summaries, start=1):
            doctor = session.get(Doctor, summary.doctor_id)
            summary_data.append(
                [
                    index,
                    _short_text(doctor.name if doctor else f"Dokter {summary.doctor_id}", 28),
                    _money(summary.treatment_fee_total),
                    _money(summary.ortho_fee_total),
                    _money(summary.total_bill),
                    _money(summary.tax),
                    _money(summary.transfer_amount),
                    doctor.bank_name if doctor and doctor.bank_name else "-",
                    doctor.account_number if doctor and doctor.account_number else "-",
                ]
            )
        summary_data.append(
            [
                "",
                "TOTAL",
                _money(sum(row.treatment_fee_total for row in all_summaries)),
                _money(sum(row.ortho_fee_total for row in all_summaries)),
                _money(sum(row.total_bill for row in all_summaries)),
                _money(sum(row.tax for row in all_summaries)),
                _money(sum(row.transfer_amount for row in all_summaries)),
                "",
                "",
            ]
        )
        table = Table(summary_data, colWidths=[10 * mm, 45 * mm, 30 * mm, 30 * mm, 32 * mm, 28 * mm, 32 * mm, 30 * mm, 36 * mm])
        table.setStyle(
            TableStyle(
                [
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("ALIGN", (2, 1), (6, -1), "RIGHT"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7FAFC")]),
                    ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#D9EAF7")),
                    ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                ]
            )
        )
        table_width, table_height = table.wrapOn(pdf, right - left, y)
        table.drawOn(pdf, left, y - table_height)

    has_drawn_page = include_summary
    for summary in summaries:
        doctor = session.get(Doctor, summary.doctor_id)
        doctor_name = doctor.name if doctor else f"Dokter {summary.doctor_id}"
        rows = session.exec(
            select(DoctorTransaction).where(DoctorTransaction.period == period, DoctorTransaction.doctor_id == summary.doctor_id)
        ).all()
        rows = sorted(rows, key=lambda item: (item.transaction_date, item.patient_name.casefold(), item.id or 0))

        doctor_index = all_summaries.index(summary) + 1
        styles = getSampleStyleSheet()
        treatment_style = styles["BodyText"]
        treatment_style.fontSize = 5.6
        treatment_style.leading = 6.2
        treatment_style.alignment = TA_CENTER
        detail_rows = []
        for transaction in rows:
            bhp, price = _transaction_amounts(session, transaction)
            detail_rows.append(
                [
                    transaction.transaction_date.strftime("%d/%m/%Y"),
                    _short_text(transaction.patient_name, 20),
                    Paragraph(_short_text(transaction.treatment_name_snapshot, 70), treatment_style),
                    _number(bhp),
                    _number(price),
                    _number(transaction.qty),
                    _number(transaction.discount_amount),
                    _number(transaction.service_amount),
                    _number(transaction.doctor_fee_amount),
                    _number(transaction.special_fee_amount),
                    _number(transaction.total_bill_amount),
                ]
            )
        total_row = ["", "", "TOTAL", "", "", "", _number(sum(row.discount_amount for row in rows)), _number(sum(row.service_amount for row in rows)), _number(summary.treatment_fee_total), _number(summary.ortho_fee_total), _number(summary.total_bill)]
        chunk_size = 18
        chunks = [detail_rows[index : index + chunk_size] for index in range(0, len(detail_rows), chunk_size)] or [[]]
        headers = ["Tanggal", "Nama Pasien", "Perawatan", "BHP", "Biaya Perawatan", "QTY", "Diskon", "Biaya Jasa", "Fee Dokter", "Fee Khusus Behel", "Total Biaya"]

        for chunk_index, chunk in enumerate(chunks):
            if has_drawn_page:
                pdf.showPage()
            has_drawn_page = True
            pdf.setFont("Helvetica", 10)
            pdf.drawCentredString(width / 2, height - 12 * mm, _clinic_name())
            pdf.setFont("Helvetica-Bold", 15)
            pdf.drawCentredString(width / 2, height - 21 * mm, f"SLIP PENDAPATAN DOKTER {doctor_name}")
            pdf.setFont("Helvetica", 8)
            pdf.drawString(left, height - 30 * mm, f"No. Surat: {_letter_number(doctor_index, period)}")
            pdf.drawString(left, height - 35 * mm, f"Periode: {_period_label(period)}")
            if export_status == "DRAFT":
                pdf.drawRightString(right, height - 30 * mm, "DRAFT")
            pdf.drawRightString(
                right,
                height - 35 * mm,
                f"Fee perawatan {_rate(doctor.normal_fee_rate if doctor else None)} | Fee ortho {_rate(doctor.ortho_fee_rate if doctor else None)}",
            )
            pdf.line(left, height - 39 * mm, right, height - 39 * mm)
            table_data = [headers, *chunk]
            if chunk_index == len(chunks) - 1:
                table_data.append(total_row)
            detail_table = Table(
                table_data,
                colWidths=[17 * mm, 26 * mm, 47 * mm, 17 * mm, 22 * mm, 9 * mm, 18 * mm, 21 * mm, 21 * mm, 26 * mm, 24 * mm],
                repeatRows=1,
            )
            style_commands = [
                ("GRID", (0, 0), (-1, -1), 0.45, colors.black),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9EAF7")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (3, 1), (-1, -1), "RIGHT"),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("FONTSIZE", (0, 0), (-1, -1), 5.8),
            ]
            group_start = 0
            while group_start < len(chunk):
                group_end = group_start
                while group_end + 1 < len(chunk) and chunk[group_end + 1][0] == chunk[group_start][0]:
                    group_end += 1
                if group_end > group_start:
                    style_commands.append(("SPAN", (0, group_start + 1), (0, group_end + 1)))
                    style_commands.append(("VALIGN", (0, group_start + 1), (0, group_end + 1), "MIDDLE"))

                patient_start = group_start
                while patient_start <= group_end:
                    patient_end = patient_start
                    while patient_end + 1 <= group_end and chunk[patient_end + 1][1] == chunk[patient_start][1]:
                        patient_end += 1
                    if patient_end > patient_start:
                        style_commands.append(("SPAN", (1, patient_start + 1), (1, patient_end + 1)))
                        style_commands.append(("VALIGN", (1, patient_start + 1), (1, patient_end + 1), "MIDDLE"))
                    patient_start = patient_end + 1
                group_start = group_end + 1
            if chunk_index == len(chunks) - 1:
                style_commands.extend(
                    [
                        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#F4B183")),
                        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                        ("SPAN", (0, -1), (1, -1)),
                    ]
                )
            detail_table.setStyle(TableStyle(style_commands))
            table_width, table_height = detail_table.wrapOn(pdf, right - left, height - 50 * mm)
            detail_table.drawOn(pdf, left, height - 45 * mm - table_height)

    pdf.save()
    stream.seek(0)
    return stream


def doctor_fee_pdf_zip(session: Session, period: str) -> BytesIO:
    summaries, _ = _doctor_fee_summaries(session, period)
    stream = BytesIO()
    with ZipFile(stream, "w", ZIP_DEFLATED) as archive:
        for index, summary in enumerate(summaries, start=1):
            doctor = session.get(Doctor, summary.doctor_id)
            doctor_name = doctor.name if doctor else f"Dokter {summary.doctor_id}"
            pdf = doctor_fee_pdf(session, period, doctor_id=summary.doctor_id, include_summary=False)
            archive.writestr(f"{index:03d}-slip-fee-{period}-{_safe_filename(doctor_name)}.pdf", pdf.getvalue())
    stream.seek(0)
    return stream


def _payroll_records(session: Session, period: str) -> list[PayrollRecord]:
    rows = session.exec(select(PayrollRecord).where(PayrollRecord.period == period)).all()
    return sorted(rows, key=lambda item: (session.get(Employee, item.employee_id).name.casefold() if session.get(Employee, item.employee_id) else str(item.employee_id)))


def _payroll_status(rows: list[PayrollRecord]) -> str:
    return "FINAL" if rows and all(row.status == PeriodStatus.LOCKED for row in rows) else "DRAFT"


def _payroll_gross(row: PayrollRecord) -> float:
    return row.base_salary + row.double_shift_fee + row.sunday_fee + row.overtime_total + row.bonus + row.position_allowance


def _payroll_deductions(row: PayrollRecord) -> float:
    return row.bpjs_deduction + row.other_deduction + row.pph21


def _payroll_overtime_rows(session: Session, period: str, employee_id: int | None = None) -> list[AttendanceRecord]:
    statement = select(AttendanceRecord).where(AttendanceRecord.period == period, AttendanceRecord.overtime_minutes > 0)
    if employee_id:
        statement = statement.where(AttendanceRecord.employee_id == employee_id)
    rows = session.exec(statement).all()
    return sorted(rows, key=lambda item: (item.employee_name_snapshot.casefold(), item.work_date, item.id or 0))


def _style_payroll_recap(ws) -> None:
    ws.freeze_panes = "A5"
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    title_fill = PatternFill("solid", fgColor="4F81BD")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    total_fill = PatternFill("solid", fgColor="D9EAF7")
    for row in (1, 2):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            cell.fill = title_fill
            cell.font = Font(bold=True, size=16 if row == 1 else 11)
            cell.alignment = Alignment(horizontal="center", vertical="center")
    for cell in ws[4]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = table_border
    for row in range(5, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            cell.border = table_border
            cell.alignment = Alignment(vertical="center")
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'
        if ws.cell(row, 2).value == "TOTAL":
            for col in range(1, ws.max_column + 1):
                ws.cell(row, col).fill = total_fill
                ws.cell(row, col).font = Font(bold=True)
    widths = [6, 24, 18, 14, 12, 14, 12, 12, 10, 10, 10, 10, 16, 16, 12, 14, 16, 14, 16, 16, 14, 14, 18, 14, 16, 20, 18, 18]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width


def _style_overtime_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    total_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = table_border
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            cell.border = table_border
            cell.alignment = Alignment(vertical="center")
        if ws.cell(row, 1).value == "TOTAL":
            for col in range(1, ws.max_column + 1):
                ws.cell(row, col).fill = total_fill
                ws.cell(row, col).font = Font(bold=True)
    ws.freeze_panes = "A2"
    _autosize(ws, 12, 34)


def _write_payroll_slip_sheet(slip, records: list[PayrollRecord], session: Session, period: str) -> None:
    dark_blue = "0B2F63"
    orange = "F97316"
    light_blue = "EAF2FF"
    total_fill = "FCE4D6"
    thin_gray = Side(style="thin", color="D9E2F3")
    slip_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    money_format = '"Rp"#,##0'
    row_cursor = 1
    for record in records:
        employee = session.get(Employee, record.employee_id)
        employee_name = employee.name if employee else str(record.employee_id)
        employee_position = employee.position if employee and employee.position else "-"
        slip.merge_cells(start_row=row_cursor, start_column=1, end_row=row_cursor, end_column=5)
        title = slip.cell(row_cursor, 1)
        title.value = f"SLIP GAJI - {_clinic_name()}"
        title.fill = PatternFill("solid", fgColor=dark_blue)
        title.font = Font(bold=True, color="FFFFFF", size=14)
        title.alignment = Alignment(horizontal="center", vertical="center")
        slip.row_dimensions[row_cursor].height = 22

        info_rows = [
            ("Nama Karyawan", employee_name),
            ("Jabatan", employee_position),
            ("Periode", _period_label(period)),
        ]
        for offset, (label, value) in enumerate(info_rows, start=2):
            label_cell = slip.cell(row_cursor + offset, 1)
            value_cell = slip.cell(row_cursor + offset, 2)
            label_cell.value = label
            value_cell.value = value
            label_cell.fill = PatternFill("solid", fgColor=light_blue)
            label_cell.font = Font(bold=True)
            for col in range(1, 6):
                cell = slip.cell(row_cursor + offset, col)
                cell.border = slip_border
                cell.alignment = Alignment(vertical="center")

        income_header_row = row_cursor + 6
        slip.merge_cells(start_row=income_header_row, start_column=1, end_row=income_header_row, end_column=5)
        income_header = slip.cell(income_header_row, 1)
        income_header.value = "PENDAPATAN"
        income_header.fill = PatternFill("solid", fgColor=orange)
        income_header.font = Font(bold=True, color="FFFFFF")
        income_header.alignment = Alignment(horizontal="center", vertical="center")

        income_rows = [
            ("Gaji Pokok", record.base_salary),
            ("Bonus", record.bonus),
            ("Tunjangan", record.position_allowance),
            ("Lembur", record.overtime_total),
            ("Masuk Hari Minggu", record.sunday_fee),
            ("Double shift (Nerus)", record.double_shift_fee),
        ]
        for index, (label, value) in enumerate(income_rows, start=income_header_row + 1):
            slip.cell(index, 1).value = label
            slip.cell(index, 2).value = value

        deduction_header_row = income_header_row + 8
        slip.merge_cells(start_row=deduction_header_row, start_column=1, end_row=deduction_header_row, end_column=5)
        deduction_header = slip.cell(deduction_header_row, 1)
        deduction_header.value = "POTONGAN"
        deduction_header.fill = PatternFill("solid", fgColor=dark_blue)
        deduction_header.font = Font(bold=True, color="FFFFFF")
        deduction_header.alignment = Alignment(horizontal="center", vertical="center")

        deduction_rows = [
            ("Keterlambatan", record.other_deduction),
            ("BPJS", record.bpjs_deduction),
            ("PPh 21", record.pph21),
        ]
        for index, (label, value) in enumerate(deduction_rows, start=deduction_header_row + 1):
            slip.cell(index, 1).value = label
            slip.cell(index, 2).value = value

        total_row = deduction_header_row + 5
        slip.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=3)
        slip.cell(total_row, 1).value = "TOTAL GAJI DITERIMA"
        slip.cell(total_row, 4).value = record.net_salary
        slip.cell(total_row, 1).fill = PatternFill("solid", fgColor=total_fill)
        slip.cell(total_row, 4).fill = PatternFill("solid", fgColor=total_fill)
        slip.cell(total_row, 1).font = Font(bold=True)
        slip.cell(total_row, 4).font = Font(bold=True)

        for row in range(row_cursor, total_row + 1):
            for col in range(1, 6):
                cell = slip.cell(row, col)
                cell.border = slip_border
                cell.alignment = Alignment(vertical="center")
                if isinstance(cell.value, (int, float)):
                    cell.number_format = money_format
                    cell.alignment = Alignment(horizontal="right", vertical="center")
        row_cursor = total_row + 8
    widths = [26, 20, 18, 20, 18]
    for index, width in enumerate(widths, start=1):
        slip.column_dimensions[get_column_letter(index)].width = width


def _draw_slip_content(pdf: canvas.Canvas, record: PayrollRecord, employee: Employee | None, period: str, y_start: float = 270 * mm) -> None:
    width, _ = A4
    left = 18 * mm
    right = width - 18 * mm
    employee_name = employee.name if employee else f"Karyawan {record.employee_id}"
    pdf.setFont("Helvetica-Bold", 15)
    pdf.drawCentredString(width / 2, y_start, f"SLIP GAJI - {_clinic_name()}")
    pdf.setFont("Helvetica", 9)
    pdf.drawString(left, y_start - 12 * mm, "Nama Karyawan")
    pdf.drawString(55 * mm, y_start - 12 * mm, employee_name)
    pdf.drawString(left, y_start - 18 * mm, "Jabatan")
    pdf.drawString(55 * mm, y_start - 18 * mm, employee.position if employee and employee.position else "-")
    pdf.drawString(left, y_start - 24 * mm, "Periode")
    pdf.drawString(55 * mm, y_start - 24 * mm, _period_label(period))

    income = [
        ("Gaji Pokok", record.base_salary),
        ("Bonus", record.bonus),
        ("Tunjangan", record.position_allowance),
        ("Lembur", record.overtime_total),
        ("Masuk Hari Minggu / Libur", record.sunday_fee),
        ("Double shift (Nerus)", record.double_shift_fee),
    ]
    deductions = [
        ("Keterlambatan / Potongan Lain", record.other_deduction),
        ("BPJS", record.bpjs_deduction),
        ("PPh 21", record.pph21),
    ]
    table_data = [["PENDAPATAN", ""], *[[label, _money(value)] for label, value in income], ["POTONGAN", ""], *[[label, _money(value)] for label, value in deductions], ["TOTAL GAJI DITERIMA", _money(record.net_salary)]]
    table = Table(table_data, colWidths=[105 * mm, 55 * mm])
    style = [
        ("GRID", (0, 0), (-1, -1), 0.45, colors.black),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9EAF7")),
        ("BACKGROUND", (0, 7), (-1, 7), colors.HexColor("#D9EAF7")),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#F4B183")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 7), (-1, 7), "Helvetica-Bold"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("ALIGN", (1, 1), (1, -1), "RIGHT"),
        ("SPAN", (0, 0), (1, 0)),
        ("SPAN", (0, 7), (1, 7)),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]
    table.setStyle(TableStyle(style))
    table_width, table_height = table.wrapOn(pdf, right - left, 180 * mm)
    table.drawOn(pdf, left, y_start - 34 * mm - table_height)
    pdf.setFont("Helvetica", 9)
    pdf.drawString(left, 35 * mm, "Disetujui oleh:")
    pdf.drawString(left, 20 * mm, "________________________")


def payroll_xlsx(session: Session, period: str) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Form Gaji Karyawan"
    records = _payroll_records(session, period)
    ws.merge_cells("A1:AB1")
    ws.merge_cells("A2:G2")
    ws["A1"] = "FORM REKAP GAJI KARYAWAN"
    ws["A2"] = f"Periode Cut Off : {_period_label(period)}"
    headers = ["No", "Nama Karyawan", "Jabatan", "Join Date", "Masa Kerja", "Gaji Pokok", "Jumlah Hari Kerja", "Nerus (double shift)", "Izin", "Sakit", "Cuti", "Alpha", "fee double shift (nerus)", "Masuk Hari Minggu", "Lembur (menit)", "Tarif Lembur (menit)", "Total Lembur", "Bonus", "Tunjangan Jabatan", "Potongan BPJS TK 2% JHT", "Potongan Lain", "PPh 21", "Total Gaji Bersih", "Pembayaran", "Nama Bank", "Nama Penerima", "no rekening", "nominal transfer"]
    ws.append([])
    ws.append(headers)
    for index, record in enumerate(records, start=1):
        employee = session.get(Employee, record.employee_id)
        ws.append(
            [
                index,
                employee.name if employee else record.employee_id,
                employee.position if employee else None,
                employee.join_date if employee else None,
                None,
                record.base_salary,
                record.working_days,
                record.double_shift_count,
                record.izin_count,
                record.sakit_count,
                record.cuti_count,
                record.alpha_count,
                record.double_shift_fee,
                record.sunday_fee,
                record.overtime_minutes,
                record.overtime_rate_per_minute,
                record.overtime_total,
                record.bonus,
                record.position_allowance,
                record.bpjs_deduction,
                record.other_deduction,
                record.pph21,
                record.net_salary,
                record.payment_method,
                record.bank_name,
                record.account_name,
                record.account_number,
                record.net_salary,
            ]
        )
    if records:
        ws.append(["", "TOTAL", "", "", "", sum(row.base_salary for row in records), "", sum(row.double_shift_count for row in records), sum(row.izin_count for row in records), sum(row.sakit_count for row in records), sum(row.cuti_count for row in records), sum(row.alpha_count for row in records), sum(row.double_shift_fee for row in records), sum(row.sunday_fee for row in records), sum(row.overtime_minutes for row in records), "", sum(row.overtime_total for row in records), sum(row.bonus for row in records), sum(row.position_allowance for row in records), sum(row.bpjs_deduction for row in records), sum(row.other_deduction for row in records), sum(row.pph21 for row in records), sum(row.net_salary for row in records), "", "", "", "", sum(row.net_salary for row in records)])
    _style_payroll_recap(ws)

    overtime = wb.create_sheet("REKAPAN LEMBUR")
    overtime.append(["Nama", "Tanggal", "Timezone I", "Timezone II", "Menit Lembur", "Catatan"])
    overtime_has_rows = False
    for record in records:
        employee = session.get(Employee, record.employee_id)
        rows = _payroll_overtime_rows(session, period, record.employee_id)
        if not rows and record.overtime_minutes <= 0:
            continue
        overtime_has_rows = True
        start = overtime.max_row + 1
        if rows:
            for row in rows:
                overtime.append([employee.name if employee else row.employee_name_snapshot, row.work_date, f"{row.timezone1_in or '-'} / {row.timezone1_out or '-'}", f"{row.timezone2_in or '-'} / {row.timezone2_out or '-'}", row.overtime_minutes, row.status_note])
            total_overtime = sum(row.overtime_minutes for row in rows)
        else:
            overtime.append([employee.name if employee else record.employee_id, "-", "-", "-", record.overtime_minutes, "Detail absensi lembur tidak tersedia."])
            total_overtime = record.overtime_minutes
        overtime.append(["TOTAL", "", "", "", total_overtime, ""])
        if overtime.max_row - 1 > start:
            overtime.merge_cells(start_row=start, start_column=1, end_row=overtime.max_row - 1, end_column=1)
    if not overtime_has_rows:
        overtime.append(["Tidak ada data lembur untuk periode ini.", "", "", "", 0, ""])
        overtime.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
    _style_overtime_sheet(overtime)

    slip = wb.create_sheet("SLIP GAJI")
    _write_payroll_slip_sheet(slip, records, session, period)
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def payroll_pdf(session: Session, period: str, employee_id: int | None = None, include_summary: bool = True) -> BytesIO:
    all_records = _payroll_records(session, period)
    records = [row for row in all_records if employee_id is None or row.employee_id == employee_id]
    export_status = _payroll_status(all_records)
    stream = BytesIO()
    pdf = canvas.Canvas(stream, pagesize=landscape(A4))
    has_page = False
    if include_summary:
        y = _draw_pdf_header(pdf, "REKAP PAYROLL KARYAWAN", period, export_status)
        summary_data = [["NO", "NAMA", "JABATAN", "GAJI POKOK", "LEMBUR", "POTONGAN", "TRANSFER", "BANK", "REKENING"]]
        for index, record in enumerate(all_records, start=1):
            employee = session.get(Employee, record.employee_id)
            summary_data.append([index, _short_text(employee.name if employee else str(record.employee_id), 26), _short_text(employee.position if employee and employee.position else "-", 18), _money(record.base_salary), _money(record.overtime_total), _money(_payroll_deductions(record)), _money(record.net_salary), record.bank_name or "-", record.account_number or "-"])
        summary_data.append(["", "TOTAL", "", _money(sum(row.base_salary for row in all_records)), _money(sum(row.overtime_total for row in all_records)), _money(sum(_payroll_deductions(row) for row in all_records)), _money(sum(row.net_salary for row in all_records)), "", ""])
        table = Table(summary_data, colWidths=[10 * mm, 42 * mm, 35 * mm, 30 * mm, 28 * mm, 28 * mm, 32 * mm, 30 * mm, 38 * mm])
        table.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), 0.5, colors.black), ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("ALIGN", (3, 1), (6, -1), "RIGHT"), ("FONTSIZE", (0, 0), (-1, -1), 7), ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#D9EAF7")), ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold")]))
        table_width, table_height = table.wrapOn(pdf, 277 * mm, y)
        table.drawOn(pdf, 10 * mm, y - table_height)
        has_page = True
    for record in records:
        if has_page:
            pdf.showPage()
        has_page = True
        pdf.setPageSize(A4)
        _draw_slip_content(pdf, record, session.get(Employee, record.employee_id), period)
    pdf.save()
    stream.seek(0)
    return stream


def payroll_slip_pdf(session: Session, period: str, employee_id: int) -> BytesIO:
    return payroll_pdf(session, period, employee_id=employee_id, include_summary=False)


def payroll_pdf_zip(session: Session, period: str) -> BytesIO:
    records = _payroll_records(session, period)
    stream = BytesIO()
    with ZipFile(stream, "w", ZIP_DEFLATED) as archive:
        for index, record in enumerate(records, start=1):
            employee = session.get(Employee, record.employee_id)
            name = employee.name if employee else f"Karyawan {record.employee_id}"
            pdf = payroll_slip_pdf(session, period, record.employee_id)
            archive.writestr(f"{index:03d}-slip-gaji-{period}-{_safe_filename(name)}.pdf", pdf.getvalue())
    stream.seek(0)
    return stream


def template_xlsx(template_name: str, session: Session | None = None) -> BytesIO:
    specs = {
        "treatments": {
            "sheet": "Treatments",
            "headers": ["code", "name", "category", "doctor_cost", "specialist_cost", "bhp_cost", "service_fee", "treatment_price", "notes"],
            "sample": ["KON-001", "Konsultasi A", "KONSULTASI", 50000, 150000, 0, 50000, 50000, "Contoh keterangan"],
        },
        "doctors": {
            "sheet": "Doctors",
            "headers": ["name", "sheet_name", "bank_name", "account_name", "account_number", "nik", "normal_fee_rate", "ortho_fee_rate", "tax_rate"],
            "sample": ["Drg. Dokter 1", "Dokter 1", "MANDIRI", "Drg. Dokter 1", "1610010012345", "5201125809840001", 0.6, 0.7, 0.025],
        },
        "employees": {
            "sheet": "Employees",
            "headers": ["attendance_id", "name", "position", "join_date", "base_salary", "working_days", "is_training", "bank_name", "account_name", "account_number"],
            "sample": ["1", "Nama Karyawan 1", "Supervisor", "2026-05-01", 2712250, 25, "tidak", "BSI", "Nama Karyawan 1", "1234567890"],
        },
        "attendance": {
            "sheet": "Attendance",
            "headers": [
                "ID",
                "Nama",
                "Tgl",
                "Timezone I Masuk",
                "Timezone I Keluar",
                "Timezone II Masuk",
                "Timezone II Keluar",
                "Libur",
                "Catatan",
            ],
            "sample": ["1", "Nama Karyawan 1", "2026-05-02", "08:00", "16:00", "", "", "", ""],
        },
        "doctor-transactions": {
            "sheet": "DoctorTransactions",
            "headers": [
                "period",
                "transaction_date",
                "doctor_name",
                "patient_name",
                "treatment_name",
                "qty",
                "discount_amount",
                "bhp_override",
                "price_override",
                "special_fee_amount",
                "fee_rate",
            ],
            "sample": ["2026-05", "2026-05-02", "Drg. Dokter 1", "Nama Pasien", "Konsultasi A", 1, 0, "", "", 0, ""],
        },
    }
    spec = specs[template_name]
    wb = Workbook()
    ws = wb.active
    ws.title = spec["sheet"]
    ws.append(spec["headers"])
    ws.append(spec["sample"])
    ws.freeze_panes = "A2"
    for column in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = min(max(max_length + 2, 12), 32)
    notes = wb.create_sheet("Notes")
    notes.append(["Rule"])
    notes.append(["Jangan isi formula Excel. Isi value final saja. Formula tanpa cached value akan ditolak importer."])
    notes.append(["Header harus tetap sama seperti template."])
    if template_name == "attendance":
        notes.append(["Kolom Libur: isi ya/true/1/libur/merah untuk libur. Isi tidak/no/0/masuk/kerja agar hari Minggu tetap dianggap hari kerja biasa. Kosong: Minggu otomatis libur."])
    if template_name == "attendance" and session:
        notes.append([])
        notes.append(["Daftar ID Absensi Karyawan"])
        notes.append(["attendance_id", "name", "status"])
        employees = session.exec(select(Employee).order_by(Employee.name)).all()
        for employee in employees:
            notes.append([
                employee.attendance_id or str(employee.id or ""),
                employee.name,
                "aktif" if employee.is_active else "nonaktif",
            ])
        for column in notes.columns:
            max_length = max(len(str(cell.value or "")) for cell in column)
            notes.column_dimensions[column[0].column_letter].width = min(max(max_length + 2, 12), 36)
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream
