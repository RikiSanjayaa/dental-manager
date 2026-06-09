from datetime import date, datetime, time
from typing import Any, TypeVar

from pathlib import Path
from uuid import uuid4

from fastapi import APIRouter, File, HTTPException, UploadFile
from openpyxl import load_workbook
from pydantic import BaseModel
from sqlalchemy.exc import IntegrityError
from sqlmodel import SQLModel, select

from app.dependencies import AdminUser, CurrentUser, SessionDep
from app.config import get_settings
from app.importers import commit_doctors, commit_employees, commit_treatments, preview_doctors, preview_employees, preview_treatments
from app.models import AttendanceHoliday, AttendanceRule, Doctor, DoctorFeeRule, Employee, ImportFile, ImportKind, ImportStatus, PayrollRule, Treatment, User, UserRole
from app.security import hash_password
from app.utils import normalize_text

router = APIRouter(tags=["master-data"])
ModelT = TypeVar("ModelT", bound=SQLModel)


class UserCreate(BaseModel):
    username: str
    full_name: str
    password: str
    role: UserRole = UserRole.OPERATOR
    is_active: bool = True


class UserRead(BaseModel):
    id: int
    username: str
    full_name: str
    role: UserRole
    is_active: bool


class UserUpdate(BaseModel):
    full_name: str | None = None
    password: str | None = None
    role: UserRole | None = None
    is_active: bool | None = None


class EmployeeInput(BaseModel):
    name: str
    attendance_id: str | None = None
    position: str | None = None
    join_date: str | None = None
    base_salary: float = 2_712_250
    working_days: int = 25
    is_training: bool = False
    bank_name: str | None = None
    account_name: str | None = None
    account_number: str | None = None
    is_active: bool = True


class DoctorInput(BaseModel):
    name: str
    bank_name: str | None = None
    account_name: str | None = None
    account_number: str | None = None
    nik: str | None = None
    normal_fee_rate: float = 0.60
    ortho_fee_rate: float = 0.70
    tax_rate: float = 0.025
    is_active: bool = True


class TreatmentInput(BaseModel):
    code: str | None = None
    name: str
    category: str | None = None
    doctor_cost: float = 0
    specialist_cost: float = 0
    bhp_cost: float = 0
    service_fee: float = 0
    treatment_price: float = 0
    notes: str | None = None
    is_active: bool = True


class PayrollRuleInput(BaseModel):
    name: str
    is_default: bool = False
    default_base_salary: float = 0
    bpjs_jht_rate: float = 0.02
    overtime_rate_per_minute: float = 250
    pph21_threshold: float = 5_400_000
    pph21_rate: float = 0.05
    sunday_multiplier: float = 6 / 7
    double_shift_multiplier: float = 1.0


class DoctorFeeRuleInput(BaseModel):
    name: str
    is_default: bool = False
    normal_fee_rate: float = 0.60
    ortho_fee_rate: float = 0.70
    tax_rate: float = 0.025
    default_deduction: float = 0


class AttendanceRuleInput(BaseModel):
    name: str
    is_default: bool = False
    timezone1_start: time = time(8, 0)
    timezone1_end: time = time(16, 0)
    timezone2_start: time = time(14, 0)
    timezone2_end: time = time(21, 0)


class AttendanceHolidayInput(BaseModel):
    holiday_date: date
    name: str | None = None
    is_holiday: bool = True


def _list(session: SessionDep, model: type[ModelT], search: str | None = None) -> list[ModelT]:
    statement = select(model)
    rows = session.exec(statement).all()
    if search:
        lowered = search.lower()
        rows = [row for row in rows if lowered in str(getattr(row, "name", getattr(row, "username", ""))).lower()]
    return rows


def _get_or_404(session: SessionDep, model: type[ModelT], item_id: int) -> ModelT:
    item = session.get(model, item_id)
    if not item:
        raise HTTPException(status_code=404, detail="Data tidak ditemukan")
    return item


def _master_model(target: str) -> type[Employee] | type[Doctor] | type[Treatment]:
    if target == "employees":
        return Employee
    if target == "doctors":
        return Doctor
    if target == "treatments":
        return Treatment
    raise HTTPException(status_code=404, detail="Target master data tidak dikenal.")


@router.get("/users", response_model=list[UserRead])
def list_users(session: SessionDep, _: AdminUser) -> list[User]:
    return session.exec(select(User)).all()


@router.post("/users", response_model=UserRead)
def create_user(payload: UserCreate, session: SessionDep, _: AdminUser) -> User:
    user = User(
        username=payload.username,
        full_name=payload.full_name,
        role=payload.role,
        hashed_password=hash_password(payload.password),
        is_active=payload.is_active,
    )
    session.add(user)
    session.commit()
    session.refresh(user)
    return user


@router.patch("/users/{item_id}", response_model=UserRead)
def update_user(item_id: int, payload: UserUpdate, session: SessionDep, _: AdminUser) -> User:
    user = _get_or_404(session, User, item_id)
    for field, value in payload.model_dump(exclude_unset=True, exclude={"password"}).items():
        setattr(user, field, value)
    if payload.password:
        user.hashed_password = hash_password(payload.password)
    session.add(user)
    session.commit()
    session.refresh(user)
    return user


@router.get("/employees", response_model=list[Employee])
def list_employees(session: SessionDep, _: CurrentUser, search: str | None = None) -> list[Employee]:
    return _list(session, Employee, search)


@router.post("/employees", response_model=Employee)
def create_employee(payload: EmployeeInput, session: SessionDep, _: AdminUser) -> Employee:
    item = Employee(**payload.model_dump())
    session.add(item)
    session.commit()
    session.refresh(item)
    if not item.attendance_id:
        item.attendance_id = str(item.id)
        session.add(item)
        session.commit()
        session.refresh(item)
    return item


@router.patch("/employees/{item_id}", response_model=Employee)
def update_employee(item_id: int, payload: EmployeeInput, session: SessionDep, _: AdminUser) -> Employee:
    item = _get_or_404(session, Employee, item_id)
    for field, value in payload.model_dump().items():
        setattr(item, field, value)
    session.add(item)
    session.commit()
    session.refresh(item)
    if not item.attendance_id:
        item.attendance_id = str(item.id)
        session.add(item)
        session.commit()
        session.refresh(item)
    return item


@router.delete("/employees/{item_id}", response_model=Employee)
def delete_employee(item_id: int, session: SessionDep, _: AdminUser) -> Employee:
    item = _get_or_404(session, Employee, item_id)
    item.is_active = False
    session.add(item)
    session.commit()
    session.refresh(item)
    return item


@router.get("/doctors", response_model=list[Doctor])
def list_doctors(session: SessionDep, _: CurrentUser, search: str | None = None) -> list[Doctor]:
    return _list(session, Doctor, search)


@router.post("/doctors", response_model=Doctor)
def create_doctor(payload: DoctorInput, session: SessionDep, _: AdminUser) -> Doctor:
    item = Doctor(**payload.model_dump())
    session.add(item)
    session.commit()
    session.refresh(item)
    return item


@router.patch("/doctors/{item_id}", response_model=Doctor)
def update_doctor(item_id: int, payload: DoctorInput, session: SessionDep, _: AdminUser) -> Doctor:
    item = _get_or_404(session, Doctor, item_id)
    for field, value in payload.model_dump().items():
        setattr(item, field, value)
    session.add(item)
    session.commit()
    session.refresh(item)
    return item


@router.delete("/doctors/{item_id}", response_model=Doctor)
def delete_doctor(item_id: int, session: SessionDep, _: AdminUser) -> Doctor:
    item = _get_or_404(session, Doctor, item_id)
    item.is_active = False
    session.add(item)
    session.commit()
    session.refresh(item)
    return item


@router.get("/treatments", response_model=list[Treatment])
def list_treatments(session: SessionDep, _: CurrentUser, search: str | None = None) -> list[Treatment]:
    return _list(session, Treatment, search)


@router.post("/treatments", response_model=Treatment)
def create_treatment(payload: TreatmentInput, session: SessionDep, _: AdminUser) -> Treatment:
    item = Treatment(**payload.model_dump())
    session.add(item)
    session.commit()
    session.refresh(item)
    return item


@router.patch("/treatments/{item_id}", response_model=Treatment)
def update_treatment(item_id: int, payload: TreatmentInput, session: SessionDep, _: AdminUser) -> Treatment:
    item = _get_or_404(session, Treatment, item_id)
    for field, value in payload.model_dump().items():
        setattr(item, field, value)
    session.add(item)
    session.commit()
    session.refresh(item)
    return item


@router.delete("/treatments/{item_id}", response_model=Treatment)
def delete_treatment(item_id: int, session: SessionDep, _: AdminUser) -> Treatment:
    item = _get_or_404(session, Treatment, item_id)
    item.is_active = False
    session.add(item)
    session.commit()
    session.refresh(item)
    return item


@router.post("/{target}/{item_id}/activate")
def activate_master_item(target: str, item_id: int, session: SessionDep, _: AdminUser) -> dict:
    model = _master_model(target)
    item = _get_or_404(session, model, item_id)
    item.is_active = True
    session.add(item)
    session.commit()
    session.refresh(item)
    return {"target": target, "id": item.id, "is_active": item.is_active}


@router.post("/{target}/{item_id}/deactivate")
def deactivate_master_item(target: str, item_id: int, session: SessionDep, _: AdminUser) -> dict:
    model = _master_model(target)
    item = _get_or_404(session, model, item_id)
    item.is_active = False
    session.add(item)
    session.commit()
    session.refresh(item)
    return {"target": target, "id": item.id, "is_active": item.is_active}


@router.delete("/{target}/{item_id}/permanent")
def permanently_delete_master_item(target: str, item_id: int, session: SessionDep, _: AdminUser) -> dict:
    model = _master_model(target)
    item = _get_or_404(session, model, item_id)
    try:
        session.delete(item)
        session.commit()
    except IntegrityError as exc:
        session.rollback()
        raise HTTPException(status_code=409, detail="Data masih dipakai oleh transaksi/perhitungan lain, jadi tidak bisa dihapus permanen.") from exc
    return {"target": target, "id": item_id, "deleted": True}


def _save_upload(file: UploadFile) -> Path:
    settings = get_settings()
    suffix = Path(file.filename or "upload.xlsx").suffix or ".xlsx"
    path = settings.upload_dir / f"{uuid4().hex}{suffix}"
    return path


MASTER_IMPORTS = {
    "treatments": {
        "sheet": "Treatments",
        "kind": ImportKind.MASTER_TREATMENTS,
        "preview": preview_treatments,
        "commit": commit_treatments,
        "data_key": "treatments",
    },
    "doctors": {
        "sheet": "Doctors",
        "kind": ImportKind.MASTER_DOCTORS,
        "preview": preview_doctors,
        "commit": commit_doctors,
        "data_key": "doctors",
    },
    "employees": {
        "sheet": "Employees",
        "kind": ImportKind.MASTER_EMPLOYEES,
        "preview": preview_employees,
        "commit": commit_employees,
        "data_key": "employees",
    },
}


def _master_config(target: str) -> dict[str, Any]:
    config = MASTER_IMPORTS.get(target)
    if not config:
        raise HTTPException(status_code=404, detail="Target master data tidak dikenal.")
    return config


def _ensure_master_template(path: Path, target: str) -> None:
    config = _master_config(target)
    try:
        workbook = load_workbook(path, read_only=True, data_only=False)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="File Excel tidak bisa dibaca.") from exc
    if config["sheet"] not in workbook.sheetnames:
        raise HTTPException(status_code=400, detail=f"Gunakan template khusus {target}. Sheet {config['sheet']} tidak ditemukan.")


def _existing_master_keys(session: SessionDep, target: str) -> dict[str, set[str]]:
    if target == "treatments":
        treatments = session.exec(select(Treatment)).all()
        return {
            "names": {normalize_text(item.name) for item in treatments},
            "codes": {normalize_text(item.code) for item in treatments if item.code},
        }
    if target == "doctors":
        return {"names": {normalize_text(item.name) for item in session.exec(select(Doctor)).all()}, "codes": set()}
    if target == "employees":
        return {"names": {normalize_text(item.name) for item in session.exec(select(Employee)).all()}, "codes": set()}
    return {"names": set(), "codes": set()}


def _annotate_master_preview(session: SessionDep, target: str, preview: dict[str, Any]) -> dict[str, Any]:
    config = _master_config(target)
    data_key = config["data_key"]
    existing = _existing_master_keys(session, target)
    seen_names: set[str] = set()
    seen_codes: set[str] = set()
    rows: list[dict[str, Any]] = []
    duplicate_errors: list[dict[str, Any]] = []
    duplicate_count = 0
    new_count = 0
    update_count = 0

    for item in preview.get("data", {}).get(data_key, []):
        row = dict(item)
        row_number = int(row.get("row") or 0)
        normalized_name = normalize_text(row.get("name"))
        normalized_code = normalize_text(row.get("code")) if target == "treatments" and row.get("code") else ""
        issues: list[str] = []

        if normalized_name in seen_names:
            issues.append("Nama duplikat di file.")
        if normalized_code and normalized_code in seen_codes:
            issues.append("Kode duplikat di file.")
        if normalized_name:
            seen_names.add(normalized_name)
        if normalized_code:
            seen_codes.add(normalized_code)

        if issues:
            status = "invalid"
            duplicate_count += 1
            for issue in issues:
                duplicate_errors.append({"sheet": config["sheet"], "row": row_number, "field": "duplicate-in-file", "message": issue})
        elif normalized_name in existing["names"] or (normalized_code and normalized_code in existing["codes"]):
            status = "update"
            update_count += 1
        else:
            status = "new"
            new_count += 1

        row["status"] = status
        row["issues"] = issues
        rows.append(row)

    errors = [*preview.get("errors", []), *duplicate_errors]
    valid_rows = new_count + update_count
    invalid_rows = preview.get("invalid_rows", 0) + duplicate_count
    warnings = [*preview.get("warnings", []), *(error["message"] for error in duplicate_errors)]
    summary = {
        **preview.get("summary", {}),
        "new": new_count,
        "update": update_count,
        "invalid": invalid_rows,
        "duplicate_in_file": duplicate_count,
    }
    commit_rows = [{key: value for key, value in row.items() if key not in {"status", "issues"}} for row in rows if row["status"] != "invalid"]

    return {
        "target": target,
        "kind": config["kind"].value,
        "valid_rows": valid_rows,
        "invalid_rows": invalid_rows,
        "warnings": warnings[:200],
        "errors": errors,
        "summary": summary,
        "rows": rows,
        "data": {data_key: commit_rows},
    }


async def _build_master_preview(session: SessionDep, user: User, target: str, file: UploadFile) -> dict[str, Any]:
    config = _master_config(target)
    path = _save_upload(file)
    path.write_bytes(await file.read())
    _ensure_master_template(path, target)
    raw_preview = config["preview"](path)
    preview = _annotate_master_preview(session, target, raw_preview)
    import_file = ImportFile(
        original_filename=file.filename or path.name,
        stored_path=str(path),
        kind=config["kind"],
        rows_valid=preview["valid_rows"],
        rows_invalid=preview["invalid_rows"],
        warnings_count=len(preview["warnings"]),
        preview_json=preview,
        errors_json=preview["errors"],
        created_by_id=user.id,
    )
    session.add(import_file)
    session.commit()
    session.refresh(import_file)
    return {"import_id": import_file.id, **preview}


def _commit_master_preview(session: SessionDep, target: str, import_id: int) -> dict[str, Any]:
    config = _master_config(target)
    import_file = session.get(ImportFile, import_id)
    if not import_file or import_file.kind != config["kind"]:
        raise HTTPException(status_code=404, detail="Preview import tidak ditemukan.")
    if import_file.status == ImportStatus.COMMITTED:
        raise HTTPException(status_code=409, detail="Import sudah di-commit.")

    preview = import_file.preview_json
    result = config["commit"](session, preview["data"][config["data_key"]])
    import_file.status = ImportStatus.COMMITTED
    import_file.committed_at = datetime.utcnow()
    session.add(import_file)
    session.commit()
    return {
        "target": target,
        **result,
        "invalid_rows": preview.get("invalid_rows", 0),
        "warnings": preview.get("warnings", [])[:20],
        "errors": preview.get("errors", [])[:20],
    }


@router.post("/master-data/import/{target}/preview")
async def preview_master_import(target: str, session: SessionDep, user: AdminUser, file: UploadFile = File(...)) -> dict:
    return await _build_master_preview(session, user, target, file)


@router.post("/master-data/import/{target}/{import_id}/commit")
def commit_master_import(target: str, import_id: int, session: SessionDep, _: AdminUser) -> dict:
    return _commit_master_preview(session, target, import_id)


@router.post("/master-data/import/treatments")
async def import_treatments(session: SessionDep, _: AdminUser, file: UploadFile = File(...)) -> dict:
    path = _save_upload(file)
    path.write_bytes(await file.read())
    preview = preview_treatments(path)
    result = commit_treatments(session, preview["data"]["treatments"])
    return {
        "target": "treatments",
        **result,
        "invalid_rows": preview["invalid_rows"],
        "warnings": preview["warnings"][:20],
        "errors": preview["errors"][:20],
    }


@router.post("/master-data/import/doctors")
async def import_doctors(session: SessionDep, _: AdminUser, file: UploadFile = File(...)) -> dict:
    path = _save_upload(file)
    path.write_bytes(await file.read())
    preview = preview_doctors(path)
    result = commit_doctors(session, preview["data"]["doctors"])
    return {
        "target": "doctors",
        **result,
        "invalid_rows": preview["invalid_rows"],
        "warnings": preview["warnings"][:20],
        "errors": preview["errors"][:20],
    }


@router.post("/master-data/import/employees")
async def import_employees(session: SessionDep, _: AdminUser, file: UploadFile = File(...)) -> dict:
    path = _save_upload(file)
    path.write_bytes(await file.read())
    preview = preview_employees(path)
    result = commit_employees(session, preview["data"]["employees"])
    return {
        "target": "employees",
        **result,
        "invalid_rows": preview["invalid_rows"],
        "warnings": preview["warnings"][:20],
        "errors": preview["errors"][:20],
    }


@router.get("/settings/payroll-rules", response_model=list[PayrollRule])
def list_payroll_rules(session: SessionDep, _: AdminUser) -> list[PayrollRule]:
    return session.exec(select(PayrollRule)).all()


@router.post("/settings/payroll-rules", response_model=PayrollRule)
def create_payroll_rule(payload: PayrollRuleInput, session: SessionDep, _: AdminUser) -> PayrollRule:
    if payload.is_default:
        for rule in session.exec(select(PayrollRule)).all():
            rule.is_default = False
            session.add(rule)
    item = PayrollRule(**payload.model_dump())
    session.add(item)
    session.commit()
    session.refresh(item)
    return item


@router.patch("/settings/payroll-rules/{item_id}", response_model=PayrollRule)
def update_payroll_rule(item_id: int, payload: PayrollRuleInput, session: SessionDep, _: AdminUser) -> PayrollRule:
    item = _get_or_404(session, PayrollRule, item_id)
    if payload.is_default:
        for rule in session.exec(select(PayrollRule)).all():
            rule.is_default = False
            session.add(rule)
    for field, value in payload.model_dump().items():
        setattr(item, field, value)
    session.add(item)
    session.commit()
    session.refresh(item)
    return item


@router.get("/settings/attendance-rules", response_model=list[AttendanceRule])
def list_attendance_rules(session: SessionDep, _: AdminUser) -> list[AttendanceRule]:
    return session.exec(select(AttendanceRule)).all()


@router.patch("/settings/attendance-rules/{item_id}", response_model=AttendanceRule)
def update_attendance_rule(item_id: int, payload: AttendanceRuleInput, session: SessionDep, _: AdminUser) -> AttendanceRule:
    item = _get_or_404(session, AttendanceRule, item_id)
    if payload.is_default:
        for rule in session.exec(select(AttendanceRule)).all():
            rule.is_default = False
            session.add(rule)
    for field, value in payload.model_dump().items():
        setattr(item, field, value)
    session.add(item)
    session.commit()
    session.refresh(item)
    return item


@router.get("/settings/attendance-holidays", response_model=list[AttendanceHoliday])
def list_attendance_holidays(
    session: SessionDep,
    _: CurrentUser,
    start: date | None = None,
    end: date | None = None,
) -> list[AttendanceHoliday]:
    statement = select(AttendanceHoliday)
    if start:
        statement = statement.where(AttendanceHoliday.holiday_date >= start)
    if end:
        statement = statement.where(AttendanceHoliday.holiday_date <= end)
    return session.exec(statement.order_by(AttendanceHoliday.holiday_date)).all()


@router.post("/settings/attendance-holidays", response_model=AttendanceHoliday)
def upsert_attendance_holiday(payload: AttendanceHolidayInput, session: SessionDep, _: AdminUser) -> AttendanceHoliday:
    item = session.exec(
        select(AttendanceHoliday).where(AttendanceHoliday.holiday_date == payload.holiday_date)
    ).first()
    if not item:
        item = AttendanceHoliday(holiday_date=payload.holiday_date)
    item.name = payload.name
    item.is_holiday = payload.is_holiday
    session.add(item)
    session.commit()
    session.refresh(item)
    return item


@router.delete("/settings/attendance-holidays/{item_id}")
def delete_attendance_holiday(item_id: int, session: SessionDep, _: AdminUser) -> dict[str, str]:
    item = _get_or_404(session, AttendanceHoliday, item_id)
    session.delete(item)
    session.commit()
    return {"status": "ok"}


@router.get("/settings/doctor-fee-rules", response_model=list[DoctorFeeRule])
def list_doctor_fee_rules(session: SessionDep, _: AdminUser) -> list[DoctorFeeRule]:
    return session.exec(select(DoctorFeeRule)).all()


@router.post("/settings/doctor-fee-rules", response_model=DoctorFeeRule)
def create_doctor_fee_rule(payload: DoctorFeeRuleInput, session: SessionDep, _: AdminUser) -> DoctorFeeRule:
    if payload.is_default:
        for rule in session.exec(select(DoctorFeeRule)).all():
            rule.is_default = False
            session.add(rule)
    item = DoctorFeeRule(**payload.model_dump())
    session.add(item)
    session.commit()
    session.refresh(item)
    return item


@router.patch("/settings/doctor-fee-rules/{item_id}", response_model=DoctorFeeRule)
def update_doctor_fee_rule(item_id: int, payload: DoctorFeeRuleInput, session: SessionDep, _: AdminUser) -> DoctorFeeRule:
    item = _get_or_404(session, DoctorFeeRule, item_id)
    if payload.is_default:
        for rule in session.exec(select(DoctorFeeRule)).all():
            rule.is_default = False
            session.add(rule)
    for field, value in payload.model_dump().items():
        setattr(item, field, value)
    session.add(item)
    session.commit()
    session.refresh(item)
    return item
