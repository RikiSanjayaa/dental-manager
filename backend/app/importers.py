from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlmodel import Session, select

from app.calculations import calculate_attendance_record, calculate_doctor_transaction
from app.models import (
    AttendanceHoliday,
    AttendanceRule,
    AttendanceRecord,
    Doctor,
    DoctorFeeRule,
    DoctorTransaction,
    Employee,
    Treatment,
)
from app.utils import money_to_float, normalize_text, parse_date, parse_period, parse_time

DEFAULT_BASE_SALARY = 2_712_250


def safe_cell_value(
    formula_ws,
    values_ws,
    row: int,
    col: int,
    errors: list[dict[str, Any]],
    context: str,
    *,
    required: bool = False,
) -> Any:
    formula_cell = formula_ws.cell(row, col)
    raw_value = formula_cell.value
    value = values_ws.cell(row, col).value if formula_cell.data_type == "f" else raw_value

    if formula_cell.data_type == "f" and value is None:
        errors.append({"sheet": formula_ws.title, "row": row, "field": context, "message": "Formula tidak punya cached value; data tidak diimport."})
        return None
    if isinstance(value, str) and value.strip().startswith("="):
        errors.append({"sheet": formula_ws.title, "row": row, "field": context, "message": "Formula text terdeteksi; data tidak diimport."})
        return None
    if isinstance(value, str) and value.strip().startswith("#"):
        errors.append({"sheet": formula_ws.title, "row": row, "field": context, "message": f"Excel error {value}; data tidak diimport."})
        return None
    if required and (value is None or value == ""):
        errors.append({"sheet": formula_ws.title, "row": row, "field": context, "message": "Data wajib kosong; baris tidak diimport."})
        return None
    return value


def safe_money(
    formula_ws,
    values_ws,
    row: int,
    col: int,
    errors: list[dict[str, Any]],
    context: str,
    *,
    required: bool = False,
    default: float | None = 0,
) -> float | None:
    value = safe_cell_value(formula_ws, values_ws, row, col, errors, context, required=required)
    if value is None or value == "":
        return default
    return money_to_float(value)


def get_sheet(wb, preferred_name: str):
    return wb[preferred_name] if preferred_name in wb.sheetnames else wb.active


def header_map(sheet, header_row: int = 1) -> dict[str, int]:
    headers: dict[str, int] = {}
    for col in range(1, sheet.max_column + 1):
        value = sheet.cell(header_row, col).value
        if value:
            headers[normalize_text(value).replace(" ", "_")] = col
    return headers


def find_header_row(sheet, required: set[str], max_scan: int = 10) -> int:
    for row in range(1, min(sheet.max_row, max_scan) + 1):
        headers = set(header_map(sheet, row).keys())
        if required.issubset(headers):
            return row
    return 1


def attendance_header_map(sheet, header_row: int) -> dict[str, int]:
    headers = header_map(sheet, header_row)
    for col in range(1, sheet.max_column + 1):
        parent = normalize_text(sheet.cell(header_row, col).value).replace(" ", "_")
        child = normalize_text(sheet.cell(header_row + 1, col).value).replace(" ", "_")
        if parent in {"timezone_i", "timezone_1"} and child in {"masuk", "keluar"}:
            headers[f"timezone_i_{child}"] = col
        if parent in {"timezone_ii", "timezone_2"} and child in {"masuk", "keluar"}:
            headers[f"timezone_ii_{child}"] = col
    return headers


def first_by_header(
    formula_ws,
    values_ws,
    headers: dict[str, int],
    row: int,
    keys: list[str],
    errors: list[dict[str, Any]],
    *,
    required: bool = False,
    field: str | None = None,
) -> Any:
    for key in keys:
        if key in headers:
            return get_by_header(formula_ws, values_ws, headers, row, key, errors, required=required)
    if required:
        errors.append({"sheet": formula_ws.title, "row": row, "field": field or keys[0], "message": "Kolom wajib tidak ditemukan."})
    return None


def truthy_cell(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    text = normalize_text(value)
    return text in {"1", "true", "yes", "ya", "y", "libur", "holiday", "merah"}


def falsey_cell(value: Any) -> bool:
    if isinstance(value, bool):
        return not value
    if isinstance(value, (int, float)):
        return value == 0
    text = normalize_text(value)
    return text in {"0", "false", "no", "tidak", "n", "masuk", "kerja", "bukan_libur", "hari_kerja"}


def holiday_from_cell(value: Any) -> bool | None:
    if value not in (None, ""):
        if truthy_cell(value):
            return True
        if falsey_cell(value):
            return False
    return None


def get_by_header(formula_ws, values_ws, headers: dict[str, int], row: int, key: str, errors: list[dict[str, Any]], *, required: bool = False) -> Any:
    col = headers.get(key)
    if not col:
        if required:
            errors.append({"sheet": formula_ws.title, "row": row, "field": key, "message": "Kolom wajib tidak ditemukan."})
        return None
    return safe_cell_value(formula_ws, values_ws, row, col, errors, key, required=required)


def money_by_header(
    formula_ws,
    values_ws,
    headers: dict[str, int],
    row: int,
    key: str,
    errors: list[dict[str, Any]],
    *,
    required: bool = False,
    default: float | None = 0,
) -> float | None:
    col = headers.get(key)
    if not col:
        if required:
            errors.append({"sheet": formula_ws.title, "row": row, "field": key, "message": "Kolom wajib tidak ditemukan."})
        return default
    return safe_money(formula_ws, values_ws, row, col, errors, key, required=required, default=default)


def preview_treatments(path: str | Path) -> dict[str, Any]:
    formula_wb = load_workbook(path, data_only=False)
    values_wb = load_workbook(path, data_only=True)
    ws = get_sheet(formula_wb, "Treatments")
    values_ws = values_wb[ws.title]
    headers = header_map(values_ws)
    errors: list[dict[str, Any]] = []
    treatments: list[dict[str, Any]] = []
    for row in range(2, ws.max_row + 1):
        before = len(errors)
        name = get_by_header(ws, values_ws, headers, row, "name", errors, required=True)
        if not name:
            continue
        bhp_cost = money_by_header(ws, values_ws, headers, row, "bhp_cost", errors, default=0) or 0
        treatment_price = money_by_header(ws, values_ws, headers, row, "treatment_price", errors, default=0) or 0
        service_fee = money_by_header(ws, values_ws, headers, row, "service_fee", errors, default=None)
        treatments.append(
            {
                "row": row,
                "code": str(get_by_header(ws, values_ws, headers, row, "code", errors) or "").strip() or None,
                "name": str(name).strip(),
                "category": str(get_by_header(ws, values_ws, headers, row, "category", errors) or "").strip() or None,
                "doctor_cost": money_by_header(ws, values_ws, headers, row, "doctor_cost", errors, default=0) or 0,
                "specialist_cost": money_by_header(ws, values_ws, headers, row, "specialist_cost", errors, default=0) or 0,
                "bhp_cost": bhp_cost,
                "service_fee": service_fee if service_fee is not None else max(treatment_price - bhp_cost, 0),
                "treatment_price": treatment_price,
                "notes": str(get_by_header(ws, values_ws, headers, row, "notes", errors) or "").strip() or None,
                "valid": len(errors) == before,
            }
        )
    valid = [item for item in treatments if item.get("valid", True)]
    return {"kind": "treatments", "valid_rows": len(valid), "invalid_rows": len(treatments) - len(valid) + len(errors), "warnings": [], "errors": errors, "summary": {"treatments": len(valid)}, "data": {"treatments": valid}}


def preview_doctors(path: str | Path) -> dict[str, Any]:
    formula_wb = load_workbook(path, data_only=False)
    values_wb = load_workbook(path, data_only=True)
    ws = get_sheet(formula_wb, "Doctors")
    values_ws = values_wb[ws.title]
    headers = header_map(values_ws)
    errors: list[dict[str, Any]] = []
    doctors: list[dict[str, Any]] = []
    for row in range(2, ws.max_row + 1):
        before = len(errors)
        name = get_by_header(ws, values_ws, headers, row, "name", errors, required=True)
        if not name:
            continue
        doctors.append(
            {
                "row": row,
                "name": str(name).strip(),
                "bank_name": str(get_by_header(ws, values_ws, headers, row, "bank_name", errors) or "").strip() or None,
                "account_name": str(get_by_header(ws, values_ws, headers, row, "account_name", errors) or name).strip(),
                "account_number": str(get_by_header(ws, values_ws, headers, row, "account_number", errors) or "").strip() or None,
                "nik": str(get_by_header(ws, values_ws, headers, row, "nik", errors) or "").strip() or None,
                "normal_fee_rate": money_by_header(ws, values_ws, headers, row, "normal_fee_rate", errors, default=0.6) or 0.6,
                "ortho_fee_rate": money_by_header(ws, values_ws, headers, row, "ortho_fee_rate", errors, default=0.7) or 0.7,
                "tax_rate": money_by_header(ws, values_ws, headers, row, "tax_rate", errors, default=0.025) or 0.025,
                "valid": len(errors) == before,
            }
        )
    valid = [item for item in doctors if item.get("valid", True)]
    return {"kind": "doctors", "valid_rows": len(valid), "invalid_rows": len(doctors) - len(valid) + len(errors), "warnings": [], "errors": errors, "summary": {"doctors": len(valid)}, "data": {"doctors": valid}}


def preview_employees(path: str | Path) -> dict[str, Any]:
    formula_wb = load_workbook(path, data_only=False)
    values_wb = load_workbook(path, data_only=True)
    ws = get_sheet(formula_wb, "Employees")
    values_ws = values_wb[ws.title]
    headers = header_map(values_ws)
    errors: list[dict[str, Any]] = []
    employees: list[dict[str, Any]] = []
    for row in range(2, ws.max_row + 1):
        before = len(errors)
        name = get_by_header(ws, values_ws, headers, row, "name", errors, required=True)
        base_salary = money_by_header(ws, values_ws, headers, row, "base_salary", errors, default=DEFAULT_BASE_SALARY) or DEFAULT_BASE_SALARY
        if not name:
            continue
        training_value = first_by_header(ws, values_ws, headers, row, ["is_training", "masa_training", "training"], errors)
        employees.append(
            {
                "row": row,
                "name": str(name).strip(),
                "attendance_id": str(get_by_header(ws, values_ws, headers, row, "attendance_id", errors) or "").strip() or None,
                "position": str(get_by_header(ws, values_ws, headers, row, "position", errors) or "").strip() or None,
                "join_date": str(get_by_header(ws, values_ws, headers, row, "join_date", errors) or "").strip() or None,
                "base_salary": base_salary,
                "working_days": int(money_by_header(ws, values_ws, headers, row, "working_days", errors, default=25) or 25),
                "is_training": truthy_cell(training_value),
                "bank_name": str(get_by_header(ws, values_ws, headers, row, "bank_name", errors) or "").strip() or None,
                "account_name": str(get_by_header(ws, values_ws, headers, row, "account_name", errors) or name).strip(),
                "account_number": str(get_by_header(ws, values_ws, headers, row, "account_number", errors) or "").strip() or None,
                "valid": len(errors) == before,
            }
        )
    valid = [item for item in employees if item.get("valid", True)]
    return {"kind": "employees", "valid_rows": len(valid), "invalid_rows": len(employees) - len(valid) + len(errors), "warnings": [], "errors": errors, "summary": {"employees": len(valid)}, "data": {"employees": valid}}


def preview_attendance(path: str | Path) -> dict[str, Any]:
    formula_wb = load_workbook(path, data_only=False)
    values_wb = load_workbook(path, data_only=True)
    ws = get_sheet(formula_wb, "Attendance")
    values_ws = values_wb[ws.title]
    header_row = find_header_row(values_ws, {"id", "nama", "tgl"})
    headers = attendance_header_map(values_ws, header_row)
    errors: list[dict[str, Any]] = []
    rows: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()
    duplicate_count = 0
    for row in range(header_row + 1, ws.max_row + 1):
        before = len(errors)
        attendance_id = first_by_header(ws, values_ws, headers, row, ["id", "attendance_id"], errors, field="attendance_id")
        name = first_by_header(ws, values_ws, headers, row, ["nama", "employee_name", "name"], errors, required=True, field="employee_name")
        work_date = parse_date(first_by_header(ws, values_ws, headers, row, ["tgl", "work_date", "tanggal"], errors, required=True, field="work_date"))
        if not attendance_id and not name and not work_date:
            continue
        if not name or not work_date:
            continue
        attendance_id_text = str(attendance_id or "").strip() or None
        key = (attendance_id_text or normalize_text(name), work_date.isoformat())
        duplicate = key in seen
        if duplicate:
            duplicate_count += 1
        seen.add(key)
        timezone1_in = first_by_header(ws, values_ws, headers, row, ["timezone_i_masuk", "timezone_1_masuk", "timezone1_in"], errors)
        timezone1_out = first_by_header(ws, values_ws, headers, row, ["timezone_i_keluar", "timezone_1_keluar", "timezone1_out"], errors)
        timezone2_in = first_by_header(ws, values_ws, headers, row, ["timezone_ii_masuk", "timezone_2_masuk", "timezone2_in"], errors)
        timezone2_out = first_by_header(ws, values_ws, headers, row, ["timezone_ii_keluar", "timezone_2_keluar", "timezone2_out"], errors)
        holiday_override = holiday_from_cell(
            first_by_header(ws, values_ws, headers, row, ["libur", "hari_libur", "tanggal_merah", "holiday"], errors)
        )
        note = first_by_header(ws, values_ws, headers, row, ["catatan", "status_note", "note"], errors)
        rows.append(
            {
                "row": row,
                "attendance_id": attendance_id_text,
                "employee_name": str(name).strip(),
                "period": str(first_by_header(ws, values_ws, headers, row, ["period"], errors) or parse_period(work_date)).strip(),
                "work_date": work_date.isoformat(),
                "timezone1_in": str(timezone1_in or "") or None,
                "timezone1_out": str(timezone1_out or "") or None,
                "timezone2_in": str(timezone2_in or "") or None,
                "timezone2_out": str(timezone2_out or "") or None,
                "holiday_override": holiday_override,
                "is_holiday": bool(holiday_override),
                "status_note": str(note or "").strip() or None,
                "employee_found": True,
                "duplicate": duplicate,
                "valid": len(errors) == before,
            }
        )
    valid = [item for item in rows if item.get("valid", True)]
    return {
        "kind": "attendance",
        "valid_rows": len(valid),
        "invalid_rows": len(rows) - len(valid) + len(errors),
        "warnings": [],
        "errors": errors,
        "summary": {"attendance": len(valid), "duplicate_in_file": duplicate_count},
        "data": {"attendance": valid},
    }


def preview_doctor_transactions(path: str | Path) -> dict[str, Any]:
    formula_wb = load_workbook(path, data_only=False)
    values_wb = load_workbook(path, data_only=True)
    ws = get_sheet(formula_wb, "DoctorTransactions")
    values_ws = values_wb[ws.title]
    headers = header_map(values_ws)
    errors: list[dict[str, Any]] = []
    rows: list[dict[str, Any]] = []
    for row in range(2, ws.max_row + 1):
        before = len(errors)
        trx_date = parse_date(get_by_header(ws, values_ws, headers, row, "transaction_date", errors, required=True))
        doctor_name = get_by_header(ws, values_ws, headers, row, "doctor_name", errors, required=True)
        treatment_name = get_by_header(ws, values_ws, headers, row, "treatment_name", errors, required=True)
        if not trx_date or not doctor_name or not treatment_name:
            continue
        rows.append(
            {
                "row": row,
                "doctor_name": str(doctor_name).strip(),
                "transaction_date": trx_date.isoformat(),
                "period": str(get_by_header(ws, values_ws, headers, row, "period", errors) or parse_period(trx_date)).strip(),
                "patient_name": str(get_by_header(ws, values_ws, headers, row, "patient_name", errors) or "Nama Pasien").strip(),
                "treatment_name": str(treatment_name).strip(),
                "qty": money_by_header(ws, values_ws, headers, row, "qty", errors, default=1) or 1,
                "discount_amount": money_by_header(ws, values_ws, headers, row, "discount_amount", errors, default=0) or 0,
                "bhp_override": money_by_header(ws, values_ws, headers, row, "bhp_override", errors, default=None),
                "price_override": money_by_header(ws, values_ws, headers, row, "price_override", errors, default=None),
                "special_fee_amount": money_by_header(ws, values_ws, headers, row, "special_fee_amount", errors, default=0) or 0,
                "fee_rate": money_by_header(ws, values_ws, headers, row, "fee_rate", errors, default=None),
                "treatment_found": True,
                "valid": len(errors) == before,
            }
        )
    valid = [item for item in rows if item.get("valid", True)]
    return {"kind": "doctor_transactions", "valid_rows": len(valid), "invalid_rows": len(rows) - len(valid) + len(errors), "warnings": [], "errors": errors, "summary": {"transactions": len(valid)}, "data": {"transactions": valid, "doctors": []}}








def commit_treatments(session: Session, treatments: list[dict[str, Any]]) -> dict[str, int]:
    created = 0
    updated = 0
    for item in treatments:
        treatment = session.exec(select(Treatment).where(Treatment.name == item["name"])).first()
        if not treatment and item.get("code"):
            treatment = session.exec(select(Treatment).where(Treatment.code == item["code"])).first()
        if not treatment:
            treatment = Treatment(name=item["name"])
            created += 1
        else:
            updated += 1
        for field in ["code", "category", "doctor_cost", "specialist_cost", "bhp_cost", "service_fee", "treatment_price", "notes"]:
            setattr(treatment, field, item.get(field))
        session.add(treatment)
    session.commit()
    return {"created": created, "updated": updated}


def commit_doctors(session: Session, doctors: list[dict[str, Any]]) -> dict[str, int]:
    created = 0
    updated = 0
    for item in doctors:
        existing = session.exec(select(Doctor).where(Doctor.name == item["name"])).first()
        upsert_doctor(session, item)
        if existing:
            updated += 1
        else:
            created += 1
    session.commit()
    return {"created": created, "updated": updated}


def commit_employees(session: Session, employees: list[dict[str, Any]]) -> dict[str, int]:
    created = 0
    updated = 0
    for item in employees:
        employee = session.exec(select(Employee).where(Employee.name == item["name"])).first()
        if not employee:
            employee = Employee(name=item["name"])
            created += 1
        else:
            updated += 1
        employee.position = item["position"]
        employee.attendance_id = item.get("attendance_id") or employee.attendance_id
        employee.join_date = parse_date(item["join_date"])
        employee.base_salary = item["base_salary"]
        employee.working_days = item["working_days"]
        employee.is_training = item.get("is_training", False)
        employee.bank_name = item["bank_name"]
        employee.account_name = item["account_name"]
        employee.account_number = item["account_number"]
        session.add(employee)
        session.flush()
        if not employee.attendance_id:
            employee.attendance_id = str(employee.id)
            session.add(employee)
    session.commit()
    return {"created": created, "updated": updated}


def commit_attendance(session: Session, attendance_rows: list[dict[str, Any]]) -> dict[str, int]:
    created = 0
    updated = 0
    employees = session.exec(select(Employee)).all()
    employee_name_map = {normalize_text(employee.name): employee for employee in employees}
    employee_id_map = {
        str(employee.attendance_id or employee.id).strip(): employee
        for employee in employees
        if employee.attendance_id or employee.id
    }
    rule = session.exec(select(AttendanceRule).where(AttendanceRule.is_default == True)).first()  # noqa: E712
    rule = rule or AttendanceRule(name="Fallback", is_default=True)
    holidays = {
        holiday.holiday_date: holiday.is_holiday
        for holiday in session.exec(select(AttendanceHoliday)).all()
    }
    for item in attendance_rows:
        work_date = parse_date(item["work_date"])
        if not work_date:
            continue
        attendance_id = str(item.get("attendance_id") or "").strip() or None
        employee = (employee_id_map.get(attendance_id) if attendance_id else None) or employee_name_map.get(normalize_text(item["employee_name"]))
        statement = select(AttendanceRecord).where(
            AttendanceRecord.period == item["period"],
            AttendanceRecord.work_date == work_date,
        )
        if employee:
            statement = statement.where(AttendanceRecord.employee_id == employee.id)
        elif attendance_id:
            statement = statement.where(AttendanceRecord.attendance_id_snapshot == attendance_id)
        else:
            statement = statement.where(AttendanceRecord.employee_name_snapshot == item["employee_name"])
        record = session.exec(statement).first()
        if record:
            updated += 1
        else:
            record = AttendanceRecord(period=item["period"], work_date=work_date, employee_name_snapshot=item["employee_name"])
            created += 1
        record.employee_id = employee.id if employee else None
        record.attendance_id_snapshot = (attendance_id or employee.attendance_id or str(employee.id)) if employee else attendance_id
        record.employee_name_snapshot = item["employee_name"]
        record.timezone1_in = parse_time(item["timezone1_in"])
        record.timezone1_out = parse_time(item["timezone1_out"])
        record.timezone2_in = parse_time(item["timezone2_in"])
        record.timezone2_out = parse_time(item["timezone2_out"])
        holiday_override = item.get("holiday_override")
        record.is_holiday = bool(
            holiday_override
            if holiday_override is not None
            else holidays.get(work_date, work_date.weekday() == 6)
        )
        record.needs_review = employee is None
        record.status_note = item["status_note"] or (None if employee else "Karyawan belum ditemukan di master.")
        calculate_attendance_record(record, rule)
        session.add(record)
    session.commit()
    return {"created": created, "updated": updated}


def commit_doctor_transactions(session: Session, preview: dict[str, Any]) -> dict[str, int]:
    created = 0
    updated = 0
    treatment_map = {normalize_text(treatment.name): treatment for treatment in session.exec(select(Treatment)).all()}
    doctor_map = {normalize_text(doctor.name): doctor for doctor in session.exec(select(Doctor)).all()}
    for item in preview["data"].get("doctors", []):
        doctor = upsert_doctor(session, item)
        doctor_map[normalize_text(doctor.name)] = doctor

    default_rule = session.exec(select(DoctorFeeRule).where(DoctorFeeRule.is_default == True)).first()  # noqa: E712
    default_rule = default_rule or DoctorFeeRule(name="Default", is_default=True)

    for item in preview["data"]["transactions"]:
        if not item.get("valid", True):
            continue
        trx_date = parse_date(item["transaction_date"])
        if not trx_date:
            continue
        doctor = doctor_map.get(normalize_text(item["doctor_name"])) or upsert_doctor(session, {"name": item["doctor_name"]})
        treatment = treatment_map.get(normalize_text(item["treatment_name"]))
        trx = DoctorTransaction(
            period=item["period"],
            transaction_date=trx_date,
            doctor_id=doctor.id,
            patient_name=item["patient_name"],
            treatment_id=treatment.id if treatment else None,
            treatment_name_snapshot=item["treatment_name"],
            qty=item["qty"],
            discount_amount=item["discount_amount"],
            bhp_override=item["bhp_override"],
            price_override=item["price_override"],
            special_fee_amount=item["special_fee_amount"],
            fee_rate=item.get("fee_rate"),
            needs_review=not bool(treatment),
            review_note=None if treatment else "Treatment belum ditemukan di master.",
        )
        calculate_doctor_transaction(trx, treatment, doctor, default_rule)
        session.add(trx)
        created += 1
    session.commit()
    return {"created": created, "updated": updated}





def upsert_doctor(session: Session, item: dict[str, Any]) -> Doctor:
    name = str(item["name"]).strip()
    doctor = session.exec(select(Doctor).where(Doctor.name == name)).first()
    if not doctor:
        doctor = Doctor(name=name)
    else:
        doctor.name = name
    doctor.bank_name = item.get("bank_name") or doctor.bank_name
    doctor.account_name = item.get("account_name") or doctor.account_name
    doctor.account_number = item.get("account_number") or doctor.account_number
    doctor.nik = item.get("nik") or doctor.nik
    session.add(doctor)
    session.flush()
    return doctor
